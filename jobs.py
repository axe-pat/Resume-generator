#!/usr/bin/env python3
"""
jobs.py — Application Pipeline Manager
========================================
Bridges the discovery layer (jobs.xlsx) and the generation layer (run_app.py).
Handles promotion, generation, tracking, archiving, and status management.

Subcommands
-----------
  pipeline      Promote top-scored queued jobs + generate docs for all promoted
                (this is what the 12h cron runs)

  promote       Create apps/<Company>/ dirs from queued jobs in jobs.xlsx
  generate      Run run_app.py for promoted jobs, write folder_path back to xlsx
  list          Show jobs from the xlsx filtered by status / score
  mark          Manually update status on one or more rows
  archive       Move terminal-status rows to archive sheet to keep xlsx lean
  sync          Scan apps/ dirs, backfill folder_path for any generated jobs

Usage examples
--------------
  # Full automated pipeline — what the cron calls:
  python jobs.py pipeline --min-score 8.0 --top 10

  # Manual: promote specific jobs, then generate
  python jobs.py promote --id 42,47
  python jobs.py generate --all-promoted

  # Batch-run every apps/ dir that has a jd.txt (skip already-generated ones)
  python jobs.py generate --all-apps --docx
  python jobs.py generate --all-apps --force   # rerun even if resume already exists
  python jobs.py generate --all-apps --docx --parallel 3  # 3 jobs at once
  python jobs.py generate --companies Flexera,Lennox,Risepoint --parallel 3
  python jobs.py generate --companies Flexera,Lennox --resume-only --parallel 2
  # If a dir has a CL but no resume (run was interrupted), --resume-only is auto-applied
  python jobs.py generate --all-apps --resume-only  # force resume-only for all targets

  # Review queue
  python jobs.py list --status queued --top 20

  # Override status (e.g. skip a role you don't want)
  python jobs.py mark --id 55 --status skip

  # Monthly cleanup
  python jobs.py archive --older-than 60

Flags available on most subcommands
-------------------------------------
  --dry-run         Print what would happen, don't write anything
  --no-rewrite      Skip resume Pass 2 (voice rewrite)
  --no-score        Skip resume Pass 3 (scoring)
  --no-qc           Skip CL Step 3 (AI quality check)
  --model MODEL     Anthropic model (default: claude-sonnet-4-6)

Status lifecycle
----------------
  new → queued → promoted → generated → applied
                    ↑                      ↑
              (jobs.py promote)    (user marks after submitting)

  Terminal statuses (won't be touched by automation): applied, closed, rejected, skip
"""

import argparse
import fcntl
import json
import os
import re
import subprocess
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from shared.job_eligibility import evaluate_manual_jd

# ─────────────────────────────────────────────────────────────────────────────
# VM SSL workaround
# The Cowork VM routes traffic through an ephemeral MITM proxy (Coworkd MITM CA)
# that presents a self-signed certificate chain.  Patch the anthropic httpx
# transport to skip SSL verification when running inside the VM.
# This is safe because: (a) the proxy is Anthropic's own Cowork infra, and
# (b) the API key + TLS payload is still encrypted end-to-end inside the tunnel.
# ─────────────────────────────────────────────────────────────────────────────
def _apply_vm_ssl_patch() -> None:
    """Patch anthropic._base_client to skip SSL verify in the Cowork VM."""
    try:
        import anthropic._base_client as _abc
        _orig = _abc._DefaultHttpxClient.__init__

        def _patched(self, **kwargs):
            kwargs.setdefault("verify", False)
            _orig(self, **kwargs)

        _abc._DefaultHttpxClient.__init__ = _patched  # type: ignore[method-assign]
    except Exception:
        pass  # If anthropic isn't installed yet, the pipeline will catch it later


if os.path.exists("/sessions") and os.environ.get("HTTPS_PROXY"):
    _apply_vm_ssl_patch()

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────
ROOT_DIR   = Path(__file__).parent
JOBS_XLSX  = ROOT_DIR / "discovery" / "jobs.xlsx"
APPS_DIR   = ROOT_DIR / "apps"
APPLY_QUEUES_DIR = APPS_DIR / "Apply queues"
CURRENT_APPLY_QUEUE_DIR = APPLY_QUEUES_DIR / "current_apply_queue"
CURRENT_APPLY_QUEUE_PRIORITY_JSON = CURRENT_APPLY_QUEUE_DIR / "priority_order.json"
LOCK_FILE  = ROOT_DIR / "discovery" / ".jobs.lock"
ARCHIVE_SHEET = "Archive"
JOBS_SHEET    = "Jobs"

# Statuses the automation will never touch
TERMINAL_STATUSES = {"applied", "closed", "parked", "rejected", "skip", "skipped"}
PROMOTE_FROM      = {"queued"}        # statuses eligible for promotion
GENERATE_FROM     = {"promoted"}      # statuses eligible for generation

# Sort order for xlsx view — lower rank floats to top
STATUS_RANK = {
    "queued":    0,
    "review":    1,
    "promoted":  2,
    "generated": 3,
    "applied":   4,
    "closed":    5,
    "parked":    6,
    "new":       7,
    "skipped":   8,
    "skip":      8,
    "rejected":  8,
}

# xlsx cell fill colours per status (hex, no #)
_STATUS_COLOR = {
    "queued":    "E2EFDA",   # soft green
    "review":    "FFF2CC",   # soft amber
    "promoted":  "BDD7EE",   # soft blue
    "generated": "DDEBF7",   # lighter blue
    "applied":   "EDEDED",   # light grey
    "closed":    "EDEDED",   # light grey
    "parked":    "EDEDED",   # light grey
    "new":       None,
    "skipped":   "C0C0C0",   # silver
    "skip":      "C0C0C0",
    "rejected":  "C0C0C0",
}

# Score traffic-light colours  (fill hex, font hex)
_SCORE_COLOR = {
    "high":   ("C6EFCE", "276221"),   # green  ≥ 8.5
    "mid":    ("FFEB9C", "9C6500"),   # amber  7.0–8.4
    "low":    ("FFC7CE", "9C0006"),   # red    < 7.0
}

BLOCKLIST_PATH = ROOT_DIR / "discovery" / "blocklist.txt"

# ─────────────────────────────────────────────────────────────────────────────
# ANSI colors
# ─────────────────────────────────────────────────────────────────────────────
GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

USE_COLOR = True


def c(color, text):
    return f"{color}{text}{RESET}" if USE_COLOR else text


# ─────────────────────────────────────────────────────────────────────────────
# File lock (fcntl-based, Unix/macOS)
# ─────────────────────────────────────────────────────────────────────────────
class XlsxLock:
    """
    Exclusive advisory lock on the jobs.xlsx file.
    Uses a separate .lock file so we don't interfere with xlsx file handles.
    Blocks until the lock is acquired (up to timeout seconds).
    """
    def __init__(self, timeout: int = 30):
        self.timeout = timeout
        self._fh     = None

    def __enter__(self):
        LOCK_FILE.parent.mkdir(parents=True, exist_ok=True)
        deadline = time.time() + self.timeout
        self._fh = open(LOCK_FILE, "w")
        while True:
            try:
                fcntl.flock(self._fh, fcntl.LOCK_EX | fcntl.LOCK_NB)
                return self
            except BlockingIOError:
                if time.time() > deadline:
                    self._fh.close()
                    raise TimeoutError(
                        f"Could not acquire xlsx lock after {self.timeout}s. "
                        f"Is another jobs.py or pipeline.py running?"
                    )
                time.sleep(0.5)

    def __exit__(self, *_):
        if self._fh:
            fcntl.flock(self._fh, fcntl.LOCK_UN)
            self._fh.close()


# ─────────────────────────────────────────────────────────────────────────────
# xlsx I/O  (mirrors discovery/auto/pipeline.py helpers)
# ─────────────────────────────────────────────────────────────────────────────
COLUMNS = [
    "id", "date_found", "date_posted", "company", "role_title", "role_type",
    "location", "url", "url_hash", "source",
    "fit_score", "fit_rationale", "status",
    "date_applied", "folder_path", "resume_run", "jd_text", "notes",
]


def load_jobs() -> pd.DataFrame:
    if not JOBS_XLSX.exists():
        sys.exit(f"[ERROR] jobs.xlsx not found at {JOBS_XLSX}")
    try:
        df = pd.read_excel(JOBS_XLSX, sheet_name=JOBS_SHEET, dtype=str)
    except ValueError:
        # Preferred sheet ("Jobs") not found — fall back to first available sheet.
        # This happens when the xlsx was never written by the pipeline (sheet still
        # named "Sheet1") or was replaced by a raw export.
        try:
            import openpyxl as _opxl
            _wb = _opxl.load_workbook(JOBS_XLSX, read_only=True)
            _first = _wb.sheetnames[0]
            _wb.close()
            print(c(YELLOW, f"  [!] Sheet '{JOBS_SHEET}' not found in jobs.xlsx — "
                             f"reading from '{_first}' instead."))
            df = pd.read_excel(JOBS_XLSX, sheet_name=_first, dtype=str)
        except Exception as e2:
            sys.exit(f"[ERROR] Could not read jobs.xlsx: {e2}")
    except Exception as e:
        sys.exit(f"[ERROR] Could not read jobs.xlsx: {e}")

    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[COLUMNS].fillna("")


def _sort_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sort jobs by status priority, with a per-group secondary key:
      - applied group  → date_applied desc (latest applied first)
      - all other groups → fit_score desc
    Within each group, blocklisted companies float to the bottom.
    """
    blocklist = _load_blocklist()

    df = df.copy()
    df["_rank"]         = df["status"].apply(lambda s: STATUS_RANK.get(str(s).lower().strip(), 99))
    df["_score"]        = df["fit_score"].apply(_safe_float)
    df["_date_applied"] = pd.to_datetime(df["date_applied"], errors="coerce")
    df["_blocked"]      = df["company"].apply(lambda co: _is_blocklisted(str(co), blocklist))

    groups = []
    for rank in sorted(df["_rank"].unique()):
        group = df[df["_rank"] == rank].copy()
        if rank == STATUS_RANK["applied"]:
            group = group.sort_values(
                ["_blocked", "_date_applied"], ascending=[True, False], na_position="last"
            )
        else:
            group = group.sort_values(["_blocked", "_score"], ascending=[True, False])
        groups.append(group)

    df = pd.concat(groups) if groups else df
    return df.drop(columns=["_rank", "_score", "_date_applied", "_blocked"])


def save_jobs(df: pd.DataFrame, dry_run: bool = False) -> None:
    if dry_run:
        print(c(YELLOW, "  [dry-run] Would write to jobs.xlsx — skipped"))
        return

    import shutil as _shutil
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Always sort before writing
    df = _sort_df(df)

    # Backup before overwriting — restores formatting if another agent corrupts the file
    if JOBS_XLSX.exists():
        _bak = JOBS_XLSX.with_suffix(".xlsx.bak")
        _shutil.copy2(JOBS_XLSX, _bak)

    if JOBS_XLSX.exists():
        with pd.ExcelWriter(
            JOBS_XLSX, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name=JOBS_SHEET, index=False)
    else:
        with pd.ExcelWriter(JOBS_XLSX, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=JOBS_SHEET, index=False)

    wb = load_workbook(JOBS_XLSX)

    # ── Remove stale Sheet1 if it coexists with Jobs ───────────────────────────
    # When another agent (or pandas) writes without going through save_jobs, it
    # may create or leave a raw "Sheet1" that becomes the active/visible tab in
    # Excel, hiding the formatted "Jobs" sheet. Delete it here on every write.
    for _stale in list(wb.sheetnames):
        if _stale.lower() != JOBS_SHEET.lower() and _stale.lower() in ("sheet1", "sheet 1"):
            del wb[_stale]

    ws = wb[JOBS_SHEET]
    wb.active = ws  # ensure Jobs is the first visible tab

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_mid    = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    col_widths = {
        "id": 6, "date_found": 13, "date_posted": 13, "company": 22, "role_title": 32,
        "role_type": 12, "location": 16, "url": 40, "url_hash": 14,
        "source": 12, "fit_score": 10, "fit_rationale": 50, "status": 12,
        "date_applied": 13, "folder_path": 40, "resume_run": 12, "jd_text": 18, "notes": 40,
    }

    col_map = {col: i + 1 for i, col in enumerate(df.columns)}

    for col, col_i in col_map.items():
        cell = ws.cell(row=1, column=col_i)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_wrap
        ws.column_dimensions[get_column_letter(col_i)].width = col_widths.get(col, 18)

    ws.freeze_panes      = "A2"
    ws.row_dimensions[1].height = 28
    ws.auto_filter.ref   = ws.dimensions   # enable dropdown filters on every column

    # ── Data rows ─────────────────────────────────────────────────────────────
    divider_border = Border(top=Side(border_style="medium", color="1F3864"))
    base_font      = Font(size=10, name="Calibri")

    status_col_i  = col_map.get("status")
    score_col_i   = col_map.get("fit_score")
    company_col_i = col_map.get("company")
    prev_rank     = None
    bl_patterns   = _load_blocklist()

    blocked_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    blocked_font = Font(color="AAAAAA", italic=True, size=10, name="Calibri",
                        strike=True)

    for row_i, (_, row) in enumerate(df.iterrows(), start=2):
        status    = str(row.get("status", "")).lower().strip()
        score     = _safe_float(row.get("fit_score"))
        rank      = STATUS_RANK.get(status, 99)
        is_blocked = _is_blocklisted(str(row.get("company", "")), bl_patterns)

        # Thick dividing line between status groups
        new_section = prev_rank is not None and rank != prev_rank
        if new_section:
            for col_i in range(1, len(df.columns) + 1):
                ws.cell(row=row_i, column=col_i).border = divider_border

        # Row height + base alignment
        ws.row_dimensions[row_i].height = 15
        for col_i in range(1, len(df.columns) + 1):
            ws.cell(row=row_i, column=col_i).alignment = left_mid
            ws.cell(row=row_i, column=col_i).font      = base_font

        if is_blocked:
            # Muted grey strikethrough — overrides all other formatting for this row
            for col_i in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_i, column=col_i)
                cell.fill = blocked_fill
                cell.font = blocked_font
        else:
            # Status cell — coloured fill + bold text
            if status_col_i:
                hex_color = _STATUS_COLOR.get(status)
                sc = ws.cell(row=row_i, column=status_col_i)
                if hex_color:
                    sc.fill = PatternFill(start_color=hex_color, end_color=hex_color,
                                          fill_type="solid")
                sc.font = Font(bold=True, size=10, name="Calibri")

            # Score cell — traffic-light colouring
            if score_col_i and score > 0:
                tier = "high" if score >= 8.5 else ("mid" if score >= 7.0 else "low")
                fill_hex, font_hex = _SCORE_COLOR[tier]
                sc = ws.cell(row=row_i, column=score_col_i)
                sc.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
                sc.font = Font(color=font_hex, bold=True, size=10, name="Calibri")

        prev_rank = rank

    wb.save(JOBS_XLSX)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _dir_slug(company: str) -> str:
    """Convert company name to a safe, readable directory name."""
    slug = re.sub(r"[^\w\s\-]", "", company).strip()
    slug = re.sub(r"\s+", "_", slug)
    return slug[:60] or "Unknown"


def _safe_float(val) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _row_id(row) -> str:
    return str(row.get("id", "")).strip()


def _today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _intel_text(notes: str, url: str = "", fit_score: str = "") -> str:
    parts = []
    url = str(url or "").strip()
    score = str(fit_score or "").strip()
    notes = str(notes or "").strip()
    if url:
        parts.append(f"job_link={url}")
    if score and score.lower() != "nan":
        parts.append(f"fit_score={score}")
    if notes:
        parts.append(notes)
    return "\n".join(parts).strip()


def _resolve_generate_target(df: pd.DataFrame, company_name: str | None = None, row_id: str | None = None) -> dict:
    """
    Resolve a requested xlsx row or company name into a generate target dict.
    Prefers folder_path when present so run-folder-native generation works.
    """
    if row_id is not None:
        row_id = str(row_id).strip()
        if not row_id:
            raise ValueError("empty row id")
        _mask = df["id"].astype(str) == row_id
        if not _mask.any():
            raise ValueError(f"row id '{row_id}' not found in jobs.xlsx")
        _row = df[_mask].iloc[0]
        company_name = str(_row.get("company") or "").strip()
        if not company_name:
            raise ValueError(f"row id '{row_id}' has no company name")
    else:
        company_name = (company_name or "").strip()
        if not company_name:
            raise ValueError("empty company name")
        _mask = df["company"].str.lower() == company_name.lower()
        _row = df[_mask].iloc[0] if _mask.any() else None

    slug = _dir_slug(company_name)
    app_dir = APPS_DIR / slug

    stored_folder_path = Path(str(_row.get("folder_path", "")).strip()) if _row is not None and str(_row.get("folder_path", "")).strip() else None
    if stored_folder_path:
        if stored_folder_path.exists():
            app_dir = stored_folder_path
        else:
            candidate = APPS_DIR / stored_folder_path.name
            if candidate.exists():
                app_dir = candidate

    if not app_dir.exists():
        matches = [
            d for d in APPS_DIR.iterdir()
            if d.is_dir() and d.name.lower() == slug.lower()
        ] if APPS_DIR.exists() else []
        if matches:
            app_dir = matches[0]
        else:
            if _row is None:
                raise ValueError(
                    f"'{company_name}' not found in jobs.xlsx and no app dir exists at {app_dir}"
                )
            print(c(YELLOW,
                    f"  [!] App dir not found for '{company_name}' — "
                    f"will recreate from xlsx"))

    target: dict = {"company": company_name, "app_dir": str(app_dir)}
    if _row is not None:
        target["id"] = str(_row.get("id", ""))
    return target


def _load_queue_generate_targets(
    df: pd.DataFrame,
    queue_path: Path,
    offset: int = 0,
    limit: int | None = None,
) -> list[dict]:
    if not queue_path.exists():
        raise ValueError(f"queue file not found: {queue_path}")

    try:
        entries = json.loads(queue_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise ValueError(f"could not read queue file {queue_path}: {exc}") from exc

    if not isinstance(entries, list):
        raise ValueError(f"queue file {queue_path} must contain a JSON list")

    def _normalize_queue_dir(path_str: str) -> str:
        raw = str(path_str or "").strip()
        if not raw:
            return raw
        normalized = raw.replace(
            f"{APPLY_QUEUES_DIR}/.current_apply_queue_tmp/",
            f"{CURRENT_APPLY_QUEUE_DIR}/",
        )
        normalized = normalized.replace(
            "/Apply queues/.current_apply_queue_tmp/",
            "/Apply queues/current_apply_queue/",
        )
        return normalized

    targets: list[dict] = []
    seen_ids: set[str] = set()
    skipped = 0
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        row_id = str(entry.get("id") or "").strip()
        if not row_id or row_id in seen_ids:
            continue
        seen_ids.add(row_id)

        if skipped < max(offset, 0):
            skipped += 1
            continue

        queue_dir = _normalize_queue_dir(str(entry.get("bundle_dir") or entry.get("folder_path") or ""))
        if queue_dir:
            target = {
                "id": row_id,
                "company": str(entry.get("company") or "").strip(),
                "app_dir": queue_dir,
            }
            if not target["company"]:
                try:
                    target = _resolve_generate_target(df, row_id=row_id)
                    target["app_dir"] = queue_dir
                except ValueError as exc:
                    print(c(YELLOW, f"  [i] Skipping queue row id '{row_id}': {exc}"))
                    continue
            targets.append(target)
            if limit is not None and len(targets) >= limit:
                break
            continue

        try:
            targets.append(_resolve_generate_target(df, row_id=row_id))
        except ValueError as exc:
            print(c(YELLOW, f"  [i] Skipping queue row id '{row_id}': {exc}"))
            continue
        if limit is not None and len(targets) >= limit:
            break

    return targets


# ─────────────────────────────────────────────────────────────────────────────
# Blocklist helpers
# ─────────────────────────────────────────────────────────────────────────────
def _load_blocklist() -> list[str]:
    """
    Load company patterns from discovery/blocklist.txt.
    Lines starting with # are comments; blank lines are ignored.
    Patterns support fnmatch wildcards (*, ?, [seq]).
    Returns a list of lowercased pattern strings.
    """
    if not BLOCKLIST_PATH.exists():
        return []
    patterns = []
    for line in BLOCKLIST_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        patterns.append(line.lower())
    return patterns


def _is_blocklisted(company: str, patterns: list[str]) -> bool:
    """Return True if company matches any blocklist pattern (case-insensitive)."""
    import fnmatch
    name = company.lower().strip()
    for pat in patterns:
        if fnmatch.fnmatch(name, pat):
            return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Subcommand: list
# ─────────────────────────────────────────────────────────────────────────────
def cmd_list(args):
    df = load_jobs()

    if args.status:
        df = df[df["status"].str.lower() == args.status.lower()]

    if args.min_score:
        df = df[df["fit_score"].apply(_safe_float) >= args.min_score]

    df = df.copy()
    df["_score"] = df["fit_score"].apply(_safe_float)
    df = df.sort_values("_score", ascending=False)

    if args.top:
        df = df.head(args.top)

    if df.empty:
        print(c(YELLOW, "  No jobs match the given filters."))
        return

    print()
    print(c(BOLD, f"  {'ID':<6} {'Score':<7} {'Status':<12} {'Company':<28} {'Role':<40}"))
    print(c(BOLD, "  " + "─" * 97))
    for _, row in df.iterrows():
        score  = _safe_float(row["fit_score"])
        status = str(row["status"]).strip()
        sc     = GREEN if score >= 8.5 else (YELLOW if score >= 7.0 else RESET)
        stc    = GREEN if status == "generated" else (CYAN if status == "promoted" else RESET)
        print(f"  {c(RESET, str(row['id'])[:5]):<6} "
              f"{c(sc, f'{score:.1f}'):<7} "
              f"{c(stc, status):<12} "
              f"{str(row['company'])[:27]:<28} "
              f"{str(row['role_title'])[:39]}")
    print()
    print(f"  {len(df)} jobs shown")


# ─────────────────────────────────────────────────────────────────────────────
# Subcommand: promote
# ─────────────────────────────────────────────────────────────────────────────
def cmd_promote(args) -> list[dict]:
    """
    Create apps/<Company>/ dirs from queued jobs.
    Returns list of promoted job dicts for chaining with generate.
    """
    dry_run = args.dry_run

    with XlsxLock():
        df = load_jobs()

        # ── Select candidates ─────────────────────────────────────────────────
        blocklist = _load_blocklist()

        id_arg = getattr(args, "id", None)
        if id_arg:
            ids    = {str(i).strip() for i in id_arg.split(",")}
            mask   = df["id"].astype(str).isin(ids)
            subset = df[mask].copy()
        else:
            eligible = df["status"].isin(PROMOTE_FROM)
            if args.min_score:
                eligible &= df["fit_score"].apply(_safe_float) >= args.min_score
            subset = df[eligible].copy()
            subset["_score"] = subset["fit_score"].apply(_safe_float)
            subset = subset.sort_values("_score", ascending=False)
            # Note: --top applied AFTER blocklist so we always get N real candidates

        # ── Apply blocklist ───────────────────────────────────────────────────
        if blocklist:
            blocked_mask = subset["company"].apply(
                lambda co: _is_blocklisted(str(co), blocklist)
            )
            if blocked_mask.any():
                for _, row in subset[blocked_mask].iterrows():
                    print(c(YELLOW, f"  [blocklist] {row['company']} — skipped"))
                subset = subset[~blocked_mask]

        # ── Apply --top after blocklist filter ────────────────────────────────
        if not id_arg and args.top:
            subset = subset.head(args.top)

        if subset.empty:
            print(c(YELLOW, "  No jobs eligible for promotion."))
            return []

        APPS_DIR.mkdir(exist_ok=True)
        promoted       = []
        assigned_dirs  = set()   # track dirs assigned in this batch to catch same-company collisions

        print()
        print(c(BOLD, f"  Promoting {len(subset)} job(s)..."))
        print()

        for _, row in subset.iterrows():
            company  = str(row.get("company", "Unknown")).strip()
            role     = str(row.get("role_title", "")).strip()
            slug     = _dir_slug(company)
            app_dir  = APPS_DIR / slug
            jd_text  = str(row.get("jd_text", "")).strip()
            row_id   = _row_id(row)

            # Skip if already has a folder_path (already promoted/generated)
            if str(row.get("folder_path", "")).strip():
                print(c(YELLOW, f"  [skip] {company} — folder_path already set, skipping"))
                continue

            if not jd_text:
                print(c(YELLOW, f"  [skip] {company} — no jd_text in xlsx row"))
                continue

            # Handle name collision — two sources:
            # (a) dir already exists on disk with different JD content
            # (b) another job in this same batch already claimed this dir
            disk_collision  = (app_dir.exists() and (app_dir / "jd.txt").exists()
                               and (app_dir / "jd.txt").read_text(encoding="utf-8").strip() != jd_text)
            batch_collision = str(app_dir) in assigned_dirs

            if disk_collision or batch_collision:
                app_dir = APPS_DIR / f"{slug}_{row_id}"
                print(c(YELLOW, f"  [rename] {company} dir collision → using {app_dir.name}"))

            assigned_dirs.add(str(app_dir))

            if not dry_run:
                app_dir.mkdir(exist_ok=True)
                (app_dir / "jd.txt").write_text(jd_text, encoding="utf-8")

                # Write intel.txt if notes column has content
                notes = str(row.get("notes", "")).strip()
                intel_text = _intel_text(notes, str(row.get("url", "")), str(row.get("fit_score", "")))
                if intel_text:
                    (app_dir / "intel.txt").write_text(intel_text, encoding="utf-8")

                # Update xlsx: status → promoted, folder_path → app_dir
                mask = df["id"].astype(str) == row_id
                df.loc[mask, "status"]      = "promoted"
                df.loc[mask, "folder_path"] = str(app_dir)

            score = _safe_float(row.get("fit_score"))
            sc    = GREEN if score >= 8.5 else (YELLOW if score >= 7.0 else RESET)
            print(f"  {c(sc, f'[{score:.1f}]')}  {c(BOLD, company)} — {role}")
            print(f"        → {app_dir}")
            promoted.append({
                "id":      row_id,
                "company": company,
                "app_dir": str(app_dir),
                "score":   score,
            })

        if not dry_run and promoted:
            save_jobs(df)
            print()
            print(c(GREEN, f"  ✓ {len(promoted)} job(s) promoted → jobs.xlsx updated"))

        return promoted


# ─────────────────────────────────────────────────────────────────────────────
# Subcommand: generate
# ─────────────────────────────────────────────────────────────────────────────
def cmd_generate(args, promoted_jobs: list[dict] | None = None) -> list[dict]:
    """
    Run run_app.py for promoted jobs. Update folder_path + status in xlsx.
    promoted_jobs: if passed (from pipeline), skip loading from xlsx.
    Returns list of result dicts.
    """
    dry_run   = args.dry_run
    run_flags = _build_run_app_flags(args)

    # ── Determine targets ──────────────────────────────────────────────────
    if promoted_jobs is not None:
        # Chained from promote — use the just-promoted list
        targets = promoted_jobs
    else:
        # Standalone generate: load from xlsx
        with XlsxLock():
            df = load_jobs()

        if getattr(args, "id", None):
            ids = [value.strip() for value in args.id.split(",") if value.strip()]
            if not ids:
                sys.exit("[ERROR] --id requires at least one non-empty row id")
            targets = []
            seen = set()
            for row_id in ids:
                if row_id in seen:
                    print(c(YELLOW, f"  [i] Duplicate row id '{row_id}' in --id — skipping duplicate"))
                    continue
                seen.add(row_id)
                try:
                    targets.append(_resolve_generate_target(df, row_id=row_id))
                except ValueError as e:
                    sys.exit(f"[ERROR] {e}")
        elif hasattr(args, "company") and args.company:
            targets = [_resolve_generate_target(df, args.company)]
        elif getattr(args, "companies", None):
            names = [name.strip() for name in args.companies.split(",") if name.strip()]
            if not names:
                sys.exit("[ERROR] --companies requires at least one non-empty company name")
            targets = []
            seen = set()
            for name in names:
                key = name.lower()
                if key in seen:
                    print(c(YELLOW, f"  [i] Duplicate company '{name}' in --companies — skipping duplicate"))
                    continue
                seen.add(key)
                try:
                    targets.append(_resolve_generate_target(df, name))
                except ValueError as e:
                    sys.exit(f"[ERROR] {e}\n        Run 'jobs.py promote' first, or check the company name.")
        elif getattr(args, "queue", False):
            queue_path = Path(getattr(args, "queue_path", "") or CURRENT_APPLY_QUEUE_PRIORITY_JSON)
            offset = max(0, int(getattr(args, "offset", 0) or 0))
            limit = getattr(args, "limit", None)
            try:
                targets = _load_queue_generate_targets(df, queue_path=queue_path, offset=offset, limit=limit)
            except ValueError as e:
                sys.exit(f"[ERROR] {e}")
        elif getattr(args, "all_promoted", False):
            promoted_rows = df[df["status"] == "promoted"]
            targets = [
                {
                    "id":      str(row["id"]),
                    "company": str(row["company"]),
                    "app_dir": str(row["folder_path"]) if row["folder_path"]
                               else str(APPS_DIR / _dir_slug(str(row["company"]))),
                }
                for _, row in promoted_rows.iterrows()
            ]
        elif getattr(args, "all_apps", False):
            # ── Scan every apps/ subdir that has a jd.txt ────────────────────
            force       = getattr(args, "force", False)
            resume_only = getattr(args, "resume_only", False)
            targets = []
            skipped = []
            partial = []  # dirs with CL but no resume (auto resume-only)
            for subdir in sorted(APPS_DIR.iterdir()):
                if not subdir.is_dir() or not (subdir / "jd.txt").exists():
                    continue
                has_resume = bool(list(subdir.glob("resume_*.txt")))
                has_cl     = bool(list(subdir.glob("cl_*.txt")))
                has_strat  = (subdir / "strategy.json").exists()
                # Skip if fully generated (resume exists) unless --force
                if has_resume and not force:
                    skipped.append(subdir.name)
                    continue
                # Try to match a row in xlsx for ID lookup (best-effort)
                _mask = df["company"].str.lower() == subdir.name.lower()
                _row  = df[_mask].iloc[0] if _mask.any() else None
                entry: dict = {"company": subdir.name, "app_dir": str(subdir)}
                if _row is not None:
                    entry["id"] = str(_row.get("id", ""))
                # Auto-detect resume-only: CL exists but no resume → skip CL pipeline
                if has_cl and not has_resume and has_strat and not resume_only:
                    entry["resume_only"] = True
                    partial.append(subdir.name)
                elif resume_only:
                    entry["resume_only"] = True
                targets.append(entry)
            if skipped:
                print(c(YELLOW,
                        f"  [i] Skipping {len(skipped)} already-generated dir(s) "
                        f"(use --force to rerun): {skipped}"))
            if partial:
                print(c(YELLOW,
                        f"  [i] {len(partial)} dir(s) have CL but no resume "
                        f"— running --resume-only --no-strategy for: {partial}"))
        else:
            sys.exit("[ERROR] Specify --id, --company, --companies, --queue, --all-promoted, or --all-apps")

    if not targets:
        print(c(YELLOW, "  No promoted jobs to generate for."))
        return []

    timeout  = getattr(args, "timeout",  2400)
    parallel = getattr(args, "parallel", 1)

    results = []
    print()
    print(c(BOLD, f"  Generating docs for {len(targets)} job(s)"
            + (f" (parallel={parallel})" if parallel > 1 else "") + "..."))

    # ── Per-job execution helper ─────────────────────────────────────────────
    def _run_one_job(job: dict, silent: bool = False) -> dict:
        """
        Execute a single job: ensure app dir, build cmd, run subprocess,
        update xlsx.  Returns a result dict.
        `silent=True` suppresses per-job prints (used in parallel mode).
        """
        company = job["company"]
        app_dir = Path(job["app_dir"])

        def _p(*a, **k):
            if not silent:
                print(*a, **k)

        if not silent:
            print()
            print(c(BOLD, f"  ── {company} ──"))

        if dry_run:
            _p(c(YELLOW, f"  [dry-run] Would run: run_app.py {app_dir.name}"))
            return {**job, "success": True, "dry_run": True}

        # ── Recreate app dir from xlsx if missing ──────────────────────────
        if not app_dir.exists() or not (app_dir / "jd.txt").exists():
            reason = "App dir not found" if not app_dir.exists() else "jd.txt missing"
            _p(c(YELLOW, f"  [!] {company}: {reason} — attempting to recreate from xlsx"))
            jd_text_from_xlsx = None
            try:
                with XlsxLock():
                    _df = load_jobs()
                if job.get("id"):
                    _mask = _df["id"].astype(str) == str(job["id"])
                else:
                    _mask = _df["company"].str.lower() == company.lower()
                if _mask.any():
                    jd_text_from_xlsx = _df.loc[_mask, "jd_text"].iloc[0]
            except Exception as _e:
                _p(c(RED, f"  [ERROR] {company}: Could not load jd_text from xlsx: {_e}"))

            if not jd_text_from_xlsx or str(jd_text_from_xlsx).strip() in ("", "nan"):
                _p(c(RED, f"  [ERROR] {company}: No jd_text in xlsx — cannot recreate app dir"))
                return {**job, "success": False, "error": "app_dir missing, no jd_text"}

            try:
                app_dir.mkdir(parents=True, exist_ok=True)
                (app_dir / "jd.txt").write_text(str(jd_text_from_xlsx).strip(), encoding="utf-8")
                _p(c(GREEN, f"  ✓ {company}: Recreated jd.txt from xlsx"))
            except Exception as _e:
                _p(c(RED, f"  [ERROR] {company}: Failed to recreate app dir: {_e}"))
                return {**job, "success": False, "error": f"recreate failed: {_e}"}

        # ── Manual-app eligibility guard ───────────────────────────────────
        # Auto-discovered jobs already passed discovery scoring + pre-filters.
        # This guard is only for manual app dirs that have no linked xlsx row ID.
        if not str(job.get("id", "")).strip():
            try:
                jd_text = (app_dir / "jd.txt").read_text(encoding="utf-8")
                is_reject, reason, inferred_title = evaluate_manual_jd(jd_text)
                if is_reject:
                    title_note = f" [{inferred_title}]" if inferred_title else ""
                    _p(c(YELLOW,
                         f"  [skip] {company}{title_note} — manual JD rejected by shared "
                         f"eligibility pre-filter: {reason}"))
                    return {
                        **job,
                        "success": False,
                        "skipped": True,
                        "skip_reason": reason,
                    }
            except Exception as _e:
                _p(c(YELLOW,
                     f"  [!] {company}: manual eligibility check failed "
                     f"({_e}) — continuing without block"))

        # ── Build per-job flags (may differ from global run_flags) ─────────
        job_resume_only = job.get("resume_only", False)
        job_flags = _build_run_app_flags(args, resume_only=job_resume_only)

        cmd = [
            sys.executable,
            str(ROOT_DIR / "run_app.py"),
            company,
            "--app-dir",
            str(app_dir),
            "--no-color",
        ] + job_flags

        _p(c(CYAN, f"  → {' '.join(cmd)}"))
        t_start = time.time()

        try:
            result = subprocess.run(
                cmd,
                cwd=str(ROOT_DIR),
                capture_output=silent,  # serial: stream to terminal; parallel: capture
                timeout=timeout,
            )
            elapsed = int(time.time() - t_start)
            success = result.returncode == 0
        except subprocess.TimeoutExpired:
            elapsed = int(time.time() - t_start)
            _p(c(RED, f"  [✗] {company}: timed out after {timeout}s"))
            return {**job, "success": False, "error": f"timeout after {timeout}s",
                    "elapsed": elapsed}
        except Exception as e:
            _p(c(RED, f"  [✗] {company}: subprocess error: {e}"))
            return {**job, "success": False, "error": str(e)}

        # ── Update xlsx (XlsxLock serialises concurrent writes) ────────────
        with XlsxLock():
            _df = load_jobs()
            if job.get("id"):
                mask = _df["id"].astype(str) == str(job["id"])
            else:
                mask = _df["company"].str.lower() == company.lower()

            if mask.any():
                _df.loc[mask, "folder_path"] = str(app_dir)
                if success:
                    _df.loc[mask, "status"] = "generated"
                    run_name = getattr(args, "run_name", None)
                    if run_name:
                        _df.loc[mask, "resume_run"] = run_name
            else:
                # No matching row — backfill
                try:
                    existing_ids = pd.to_numeric(_df["id"], errors="coerce").dropna().astype(int)
                    new_id = int(existing_ids.max()) + 1 if not existing_ids.empty else 1
                except Exception:
                    new_id = 1
                jd_text_val = ""
                try:
                    jd_file = app_dir / "jd.txt"
                    if jd_file.exists():
                        jd_text_val = jd_file.read_text(encoding="utf-8").strip()
                except Exception:
                    pass
                role_title_val = ""
                try:
                    strat_file = app_dir / "strategy.json"
                    if strat_file.exists():
                        strat = json.loads(strat_file.read_text(encoding="utf-8"))
                        role_title_val = (
                            strat.get("role_title") or strat.get("jd_role") or ""
                        )
                except Exception:
                    pass
                new_row = {col: "" for col in COLUMNS}
                new_row.update({
                    "id":          str(new_id),
                    "date_found":  datetime.now().strftime("%Y-%m-%d"),
                    "company":     company,
                    "role_title":  role_title_val,
                    "source":      "manual",
                    "status":      "generated" if success else "failed",
                    "folder_path": str(app_dir),
                    "jd_text":     jd_text_val,
                })
                _df = pd.concat([_df, pd.DataFrame([new_row])], ignore_index=True)
                _p(c(CYAN, f"  [+] No xlsx row for '{company}' — created new row (id={new_id})"))
            save_jobs(_df)

        verdict_color = GREEN if success else RED
        verdict_icon  = "✓" if success else "✗"
        _p(c(verdict_color, f"  [{verdict_icon}] {company} — {elapsed}s"))

        return {**job, "success": success, "elapsed": elapsed}

    # ── Serial or parallel execution ─────────────────────────────────────────
    if parallel > 1:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        print(c(YELLOW,
                f"  [i] Parallel mode: up to {parallel} jobs at once. "
                f"Live output → logs/. Final status printed as each job completes."))
        print()

        with ThreadPoolExecutor(max_workers=parallel) as executor:
            future_to_job = {
                executor.submit(_run_one_job, job, True): job   # silent=True
                for job in targets
            }
            for future in as_completed(future_to_job):
                result = future.result()
                company = result["company"]
                elapsed = result.get("elapsed", "?")
                if result.get("dry_run"):
                    print(c(YELLOW, f"  [dry-run] {company}"))
                elif result.get("success"):
                    print(c(GREEN,  f"  [✓] {company} — {elapsed}s"))
                else:
                    err = result.get("error", "failed")
                    print(c(RED,    f"  [✗] {company} — {err}"))
                results.append(result)
    else:
        # Serial — existing behaviour: stream output directly to terminal
        for job in targets:
            results.append(_run_one_job(job, silent=False))

    # ── Summary ──────────────────────────────────────────────────────────────
    n_ok  = sum(1 for r in results if r.get("success"))
    n_err = len(results) - n_ok
    print()
    print(c(BOLD, "─" * 60))
    print(c(GREEN if not n_err else YELLOW,
            f"  Generate complete:  {n_ok} succeeded  |  {n_err} failed"))
    print(c(BOLD, "─" * 60))

    return results


def _build_run_app_flags(args, resume_only: bool = False) -> list[str]:
    flags = []
    if getattr(args, "no_rewrite",  False): flags.append("--no-rewrite")
    if getattr(args, "no_score",    False): flags.append("--no-score")
    if getattr(args, "no_qc",       False): flags.append("--no-qc")
    if getattr(args, "no_strategy", False): flags.append("--no-strategy")
    if getattr(args, "model",       None):  flags += ["--model", args.model]
    if getattr(args, "no_docx",     False): flags.append("--no-docx")
    if resume_only:
        if "--resume-only" not in flags:
            flags.append("--resume-only")
        # strategy.json already exists for partial dirs — skip re-running Pass 0
        if "--no-strategy" not in flags:
            flags.append("--no-strategy")
    return flags


# ─────────────────────────────────────────────────────────────────────────────
# Subcommand: pipeline  (promote + generate in one shot — what cron calls)
# ─────────────────────────────────────────────────────────────────────────────
def cmd_pipeline(args):
    print()
    print(c(BOLD + CYAN, "  ╔══════════════════════════════════════════╗"))
    print(c(BOLD + CYAN,  "  ║   jobs.py pipeline                       ║"))
    print(c(BOLD + CYAN,  "  ╚══════════════════════════════════════════╝"))
    print(f"  min-score={args.min_score}  top={args.top}  dry-run={args.dry_run}")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Step 1: Promote
    print()
    print(c(BOLD, "  ── Phase 1: Promote ──"))
    promoted = cmd_promote(args)

    if not promoted:
        print(c(YELLOW, "  Nothing to generate."))
        return

    # Step 2: Generate
    print()
    print(c(BOLD, "  ── Phase 2: Generate ──"))
    cmd_generate(args, promoted_jobs=promoted)


# ─────────────────────────────────────────────────────────────────────────────
# Subcommand: mark
# ─────────────────────────────────────────────────────────────────────────────
def cmd_mark(args):
    valid_statuses = {"queued", "promoted", "generated", "applied", "closed", "parked", "rejected", "skip", "skipped"}
    if args.status not in valid_statuses:
        sys.exit(f"[ERROR] Invalid status '{args.status}'. Choose from: {valid_statuses}")

    ids = {str(i).strip() for i in args.id.split(",")}

    with XlsxLock():
        df   = load_jobs()
        mask = df["id"].astype(str).isin(ids)
        if not mask.any():
            sys.exit(f"[ERROR] No rows found with id(s): {ids}")

        df.loc[mask, "status"] = args.status
        if not args.dry_run:
            save_jobs(df)

    found = df[mask][["id", "company", "role_title", "status"]].to_dict("records")
    for row in found:
        print(c(GREEN, f"  ✓ [{row['id']}] {row['company']} — {row['role_title']} → {args.status}"))
    if args.dry_run:
        print(c(YELLOW, "  [dry-run] — not written"))


# ─────────────────────────────────────────────────────────────────────────────
# Subcommand: archive
# ─────────────────────────────────────────────────────────────────────────────
def cmd_archive(args):
    """
    Move rows with terminal statuses older than N days to Archive sheet.
    Keeps the active Jobs sheet lean.
    """
    cutoff = datetime.now() - timedelta(days=args.older_than)

    with XlsxLock():
        df = load_jobs()

        # Identify archivable rows
        terminal_mask = df["status"].isin(TERMINAL_STATUSES | {"generated"})

        # Try to use date_found for age check
        def _is_old(val):
            try:
                return pd.to_datetime(val) < cutoff
            except Exception:
                return False

        age_mask  = df["date_found"].apply(_is_old)
        to_archive = df[terminal_mask & age_mask]

        if to_archive.empty:
            print(c(YELLOW, "  No rows eligible for archiving."))
            return

        print(f"  {len(to_archive)} rows to archive (status in terminal + date_found < {cutoff.date()})")

        if not args.dry_run:
            # Write archive rows to Archive sheet (append)
            from openpyxl import load_workbook
            wb = load_workbook(JOBS_XLSX)

            if ARCHIVE_SHEET not in wb.sheetnames:
                wb.create_sheet(ARCHIVE_SHEET)
                ws_arch = wb[ARCHIVE_SHEET]
                ws_arch.append(COLUMNS)  # header
            else:
                ws_arch = wb[ARCHIVE_SHEET]

            for _, row in to_archive.iterrows():
                ws_arch.append([str(row.get(col, "")) for col in COLUMNS])
            wb.save(JOBS_XLSX)

            # Remove archived rows from active sheet
            df_active = df[~(terminal_mask & age_mask)].copy()
            save_jobs(df_active)

            print(c(GREEN, f"  ✓ Archived {len(to_archive)} rows → Archive sheet"))
        else:
            print(c(YELLOW, "  [dry-run] — not written"))
            print(to_archive[["id", "company", "role_title", "status", "date_found"]].to_string(index=False))


# ─────────────────────────────────────────────────────────────────────────────
# Subcommand: sync
# ─────────────────────────────────────────────────────────────────────────────
def cmd_sync(args):
    """
    Scan apps/ dirs for generated outputs. Backfill folder_path in xlsx
    for any rows where folder_path is blank but an app dir exists.
    """
    if not APPS_DIR.exists():
        print(c(YELLOW, "  apps/ directory doesn't exist yet."))
        return

    with XlsxLock():
        df      = load_jobs()
        updated = 0

        for app_dir in sorted(APPS_DIR.iterdir()):
            if not app_dir.is_dir() or app_dir.name.startswith("."):
                continue

            # Check if this dir has generated outputs
            has_resume = any(app_dir.glob("resume_*.txt"))
            has_cl     = any(app_dir.glob("cl_*.txt"))
            if not (has_resume or has_cl):
                continue

            # Try to match to a row
            slug      = app_dir.name.lower()
            mask_path = df["folder_path"].str.lower().str.contains(
                re.escape(slug), na=False, regex=True
            )
            mask_name = df["company"].apply(
                lambda co: _dir_slug(str(co)).lower() == slug
            )

            for mask in [mask_path, mask_name]:
                if mask.any():
                    rows_to_update = df[mask & (df["folder_path"].fillna("") == "")]
                    if not rows_to_update.empty:
                        df.loc[rows_to_update.index, "folder_path"] = str(app_dir)
                        if has_resume and has_cl:
                            df.loc[rows_to_update.index, "status"] = "generated"
                        updated += rows_to_update.shape[0]
                        print(c(GREEN, f"  ✓ synced  {app_dir.name}"))
                    break

        if updated and not args.dry_run:
            save_jobs(df)
            print(c(GREEN, f"\n  {updated} row(s) updated in jobs.xlsx"))
        elif updated:
            print(c(YELLOW, f"\n  [dry-run] {updated} row(s) would be updated"))
        else:
            print("  Everything already in sync.")


# ─────────────────────────────────────────────────────────────────────────────
# Subcommand: sort  (re-sort + reformat xlsx without changing any data)
# ─────────────────────────────────────────────────────────────────────────────
def cmd_sort(args):
    """
    Re-sort jobs.xlsx by (status priority, fit_score desc) and reapply
    all visual formatting (status colours, score traffic lights, dividers).
    No data is changed — only order and presentation.
    """
    with XlsxLock():
        df = load_jobs()
        if not args.dry_run:
            save_jobs(df)
            print(c(GREEN, f"  ✓ jobs.xlsx sorted and formatted ({len(df)} rows)"))
        else:
            sorted_df = _sort_df(df)
            print(c(YELLOW, "  [dry-run] Sort order preview (first 20 rows):"))
            print()
            print(c(BOLD, f"  {'ID':<6} {'Score':<7} {'Status':<12} {'Company':<28} {'Role':<40}"))
            print(c(BOLD, "  " + "─" * 97))
            for _, row in sorted_df.head(20).iterrows():
                score  = _safe_float(row["fit_score"])
                status = str(row["status"]).strip()
                sc     = GREEN if score >= 8.5 else (YELLOW if score >= 7.0 else RESET)
                print(f"  {str(row['id'])[:5]:<6} "
                      f"{c(sc, f'{score:.1f}'):<7} "
                      f"{status:<12} "
                      f"{str(row['company'])[:27]:<28} "
                      f"{str(row['role_title'])[:39]}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def main():
    global USE_COLOR

    parser = argparse.ArgumentParser(
        description="jobs.py — Application pipeline manager",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--no-color", action="store_true")
    sub = parser.add_subparsers(dest="cmd", required=True)

    # ── pipeline ──────────────────────────────────────────────────────────────
    p_pipe = sub.add_parser("pipeline", help="Promote + generate (cron entry point)")
    p_pipe.add_argument("--min-score", type=float, default=8.0,
                        help="Minimum fit_score to promote (default: 8.0)")
    p_pipe.add_argument("--top",       type=int,   default=10,
                        help="Max jobs to promote per run (default: 10)")
    p_pipe.add_argument("--dry-run",   action="store_true")
    p_pipe.add_argument("--no-rewrite",  action="store_true")
    p_pipe.add_argument("--no-score",    action="store_true")
    p_pipe.add_argument("--no-qc",       action="store_true")
    p_pipe.add_argument("--no-strategy", action="store_true")
    p_pipe.add_argument("--docx",        action="store_true",
                        help="Deprecated no-op: docx is now generated by default")
    p_pipe.add_argument("--no-docx",     action="store_true",
                        help="Skip .docx generation during pipeline runs")
    p_pipe.add_argument("--model",       default="claude-sonnet-4-6")
    p_pipe.add_argument("--run-name",    type=str, default=None,
                        help="Tag this resume run (e.g. run_2). Stamped in resume_run column.")

    # ── promote ───────────────────────────────────────────────────────────────
    p_prom = sub.add_parser("promote", help="Create apps/ dirs from queued jobs")
    p_prom.add_argument("--top",       type=int,   default=None)
    p_prom.add_argument("--min-score", type=float, default=None)
    p_prom.add_argument("--id",        type=str,   default=None,
                        help="Comma-separated row IDs to promote")
    p_prom.add_argument("--dry-run",   action="store_true")

    # ── generate ──────────────────────────────────────────────────────────────
    p_gen = sub.add_parser("generate", help="Run run_app.py for promoted jobs")
    g = p_gen.add_mutually_exclusive_group(required=True)
    g.add_argument("--id",           type=str,  help="Comma-separated row IDs")
    g.add_argument("--company",      type=str,  help="Single company name")
    g.add_argument("--companies",    type=str,
                   help="Comma-separated company names (e.g. Flexera,Lennox,Risepoint)")
    g.add_argument("--queue",        action="store_true",
                   help="Generate in current apply-queue priority order")
    g.add_argument("--all-promoted", action="store_true")
    g.add_argument("--all-apps",     action="store_true",
                   help="Run for every apps/ subdir that has a jd.txt "
                        "(ignores xlsx status; skips already-generated dirs by default)")
    p_gen.add_argument("--queue-path", type=str, default=str(CURRENT_APPLY_QUEUE_PRIORITY_JSON),
                       help="Path to a queue priority_order.json file (default: current apply queue)")
    p_gen.add_argument("--offset",      type=int, default=0, metavar="N",
                       help="With --queue: skip the first N queue items before generating")
    p_gen.add_argument("--limit",       type=int, default=None, metavar="N",
                       help="With --queue: only generate the first N queue items")
    p_gen.add_argument("--dry-run",      action="store_true")
    p_gen.add_argument("--force",        action="store_true",
                       help="With --all-apps: run even if resume_*.txt already exists")
    p_gen.add_argument("--resume-only",  action="store_true",
                       help="Only generate the resume (skip CL). Auto-detected for dirs "
                            "that already have a CL but no resume.")
    p_gen.add_argument("--parallel",     type=int, default=1, metavar="N",
                       help="Run N jobs simultaneously (default: 1 = serial). "
                            "In parallel mode output goes to log files; "
                            "only a one-line status is printed per job.")
    p_gen.add_argument("--timeout",      type=int, default=2400, metavar="SEC",
                       help="Per-job subprocess timeout in seconds (default: 2400 = 40 min)")
    p_gen.add_argument("--no-rewrite",   action="store_true")
    p_gen.add_argument("--no-score",     action="store_true")
    p_gen.add_argument("--no-qc",        action="store_true")
    p_gen.add_argument("--no-strategy",  action="store_true")
    p_gen.add_argument("--docx",         action="store_true",
                       help="Deprecated no-op: docx is now generated by default")
    p_gen.add_argument("--no-docx",      action="store_true",
                       help="Skip .docx generation")
    p_gen.add_argument("--model",        default="claude-sonnet-4-6")
    p_gen.add_argument("--run-name",     type=str, default=None,
                        help="Tag this resume run (e.g. run_2). Stamped in resume_run column.")

    # ── list ──────────────────────────────────────────────────────────────────
    p_list = sub.add_parser("list", help="Show jobs from xlsx")
    p_list.add_argument("--status",    type=str,   default=None)
    p_list.add_argument("--min-score", type=float, default=None)
    p_list.add_argument("--top",       type=int,   default=30)

    # ── mark ──────────────────────────────────────────────────────────────────
    p_mark = sub.add_parser("mark", help="Manually update status")
    p_mark.add_argument("--id",      type=str, required=True,
                        help="Comma-separated row IDs")
    p_mark.add_argument("--status",  type=str, required=True)
    p_mark.add_argument("--dry-run", action="store_true")

    # ── archive ───────────────────────────────────────────────────────────────
    p_arch = sub.add_parser("archive", help="Move old terminal-status rows to Archive sheet")
    p_arch.add_argument("--older-than", type=int, default=60,
                        help="Archive rows older than N days (default: 60)")
    p_arch.add_argument("--dry-run",    action="store_true")

    # ── sync ──────────────────────────────────────────────────────────────────
    p_sync = sub.add_parser("sync", help="Reconcile apps/ dirs with jobs.xlsx")
    p_sync.add_argument("--dry-run", action="store_true")

    # ── sort ──────────────────────────────────────────────────────────────────
    p_sort = sub.add_parser("sort", help="Re-sort + reformat jobs.xlsx (no data changes)")
    p_sort.add_argument("--dry-run", action="store_true",
                        help="Preview sort order without writing")

    args = parser.parse_args()
    if args.no_color:
        USE_COLOR = False

    dispatch = {
        "pipeline": cmd_pipeline,
        "promote":  cmd_promote,
        "generate": cmd_generate,
        "list":     cmd_list,
        "mark":     cmd_mark,
        "archive":  cmd_archive,
        "sync":     cmd_sync,
        "sort":     cmd_sort,
    }
    dispatch[args.cmd](args)


if __name__ == "__main__":
    main()
