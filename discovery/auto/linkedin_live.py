"""
linkedin_live.py — Logged-in LinkedIn job discovery
---------------------------------------------------
Uses a real logged-in Chrome session via Playwright CDP to search LinkedIn Jobs,
extract visible job cards, open each card, capture JD text, score the results,
and append them into discovery/jobs.xlsx.

This is intentionally separate from Outreach. We borrow the same browser/session
pattern, but keep discovery logic native to this repo.

Usage:
    python discovery/auto/linkedin_live.py --dry-run
    python discovery/auto/linkedin_live.py --limit-per-search 12
    python discovery/auto/linkedin_live.py --pages 2
    python discovery/auto/linkedin_live.py --search "Product Manager Intern" --time r86400

Pre-reqs:
  1. Chrome must already be running with remote debugging enabled.
  2. The logged-in LinkedIn session must be active in that Chrome profile.
  3. Playwright must be installed in the Python env running this script.
"""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import shutil
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import quote_plus
from urllib.parse import urlencode

import pandas as pd

try:
    from playwright.sync_api import Error as PlaywrightError
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import Page, sync_playwright
except ImportError:
    print("ERROR: playwright is not installed in this Python environment.")
    print('Install it with: pip install playwright && playwright install chromium')
    sys.exit(1)

_HERE = Path(__file__).parent
sys.path.insert(0, str(_HERE))

from pipeline import COLUMNS, JOBS_XLSX, LOGS_DIR, SHEET_NAME, load_jobs, save_jobs  # noqa: E402
from scorer import DEFAULT_MODEL, score_batch  # noqa: E402


DEFAULT_DEBUG_PORT = 9222
DEFAULT_LIMIT_PER_SEARCH: int | None = None
DEFAULT_PAGES: int | None = None
SOURCE_TAG = "linkedin_live_jobs_v1"
SEARCH_RESULTS_MIN_POOL_BEFORE_FALLBACK = 2
SEARCH_RESULTS_FALLBACK_COVERAGE_RATIO = 0.6
SEARCH_RESULTS_OFFSET_PAGE_SIZE = 25
SEARCH_RESULTS_MAX_STAGNANT_OFFSET_PAGES = 2
HEARTBEAT_INTERVAL_SECONDS = 15.0
JD_REPAIR_MAX_JOBS = 12
JD_CHROME_MARKERS = (
    "looking for talent?",
    "post a job",
    "linkedin corporation ©",
    "select language",
    "questions?",
    "visit our help center",
    "manage your account and privacy",
    "recommendation transparency",
)
JD_SECTION_MARKERS = (
    "about the job",
    "responsibilities",
    "qualifications",
    "minimum qualifications",
    "preferred qualifications",
    "what you'll do",
    "what you will do",
    "what you need to succeed",
    "job description",
)
DEFAULT_SEARCHES = [
    ("Product Manager Intern", "r86400"),
    ("Product Manager Intern", "r604800"),
    ("MBA Intern", "r86400"),
    ("MBA Intern", "r604800"),
]

TIME_LABELS = {
    "r86400": "past_24h",
    "r604800": "past_week",
}
PROJECT_ROOT = JOBS_XLSX.parent.parent
APPS_DIR = PROJECT_ROOT / "apps"
RUN_EXPORTS_DIR = APPS_DIR / "runs"
REVIEW_CACHE_SHEET_NAME = "ReviewCache"
TERMINAL_CACHE_DECISIONS = {"reject", "deprioritize"}
REVIEW_CACHE_COLUMNS = [
    "cache_key",
    "url_hash",
    "tc_hash",
    "url",
    "company",
    "role_title",
    "source",
    "decision",
    "category",
    "fit_score",
    "fit_rationale",
    "notes",
    "search_term",
    "time_window",
    "date_reviewed",
]


@dataclass
class LinkedInJobCard:
    search_term: str
    time_filter: str
    title: str
    company: str
    location: str
    url: str
    listed_at: str
    insight: str
    jd_text: str


@dataclass
class SearchRunResult:
    search_term: str
    time_filter: str
    cards: list[LinkedInJobCard]
    ui_observed_count: int | None
    extracted_count: int
    route_used: str
    fallback_used: bool


@dataclass
class JdRepairSummary:
    candidates: int
    attempted: int
    repaired: int
    remaining_failed: int


def _should_try_fallback(
    primary_count: int,
    primary_ui_count: int | None,
    limit_per_search: int | None,
) -> bool:
    fallback_threshold = SEARCH_RESULTS_MIN_POOL_BEFORE_FALLBACK
    if limit_per_search is not None:
        fallback_threshold = min(limit_per_search, SEARCH_RESULTS_MIN_POOL_BEFORE_FALLBACK)
    if primary_count < fallback_threshold:
        return True

    if primary_ui_count is None or primary_ui_count <= 0:
        return False

    expected_pool = primary_ui_count
    if limit_per_search is not None:
        expected_pool = min(expected_pool, limit_per_search)
    if expected_pool <= 0:
        return False

    coverage = primary_count / expected_pool
    return coverage < SEARCH_RESULTS_FALLBACK_COVERAGE_RATIO


def _hash(text: str) -> str:
    return hashlib.md5(text.strip().lower().encode()).hexdigest()


def url_hash(url: str) -> str:
    return _hash(url)


def title_company_hash(title: str, company: str) -> str:
    return _hash(f"{title}||{company}")


def _human_pause(page: Page, low_ms: int = 900, high_ms: int = 1700) -> None:
    page.wait_for_timeout(low_ms + int((high_ms - low_ms) * 0.5))


def _jobs_search_url(search_term: str, time_filter: str) -> str:
    return (
        "https://www.linkedin.com/jobs/search-results/"
        f"?keywords={quote_plus(search_term)}"
        f"&f_TPR={time_filter}"
        "&origin=SEMANTIC_SEARCH_JOB_ALERT_IN_APP_NOTIFICATION"
    )


def _jobs_search_fallback_url(search_term: str, time_filter: str) -> str:
    return (
        "https://www.linkedin.com/jobs/search/"
        f"?keywords={quote_plus(search_term)}"
        "&distance=25"
        "&geoId=103644278"
        "&sortBy=DD"
        f"&f_TPR={time_filter}"
        "&origin=JOB_SEARCH_PAGE_JOB_FILTER"
        "&refresh=true"
    )


def _with_start_param(url: str, start: int) -> str:
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}start={start}"


def _ensure_logs_dir() -> Path:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    return LOGS_DIR


def _write_run_artifact(name: str, payload: dict) -> Path:
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    path = _ensure_logs_dir() / f"{name}_{stamp}.json"
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _write_json_artifact(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _inflight_artifact_paths(run_stamp: str, run_label: str) -> tuple[Path, Path]:
    logs_dir = _ensure_logs_dir()
    slug = _slug(run_label)
    progress_path = logs_dir / f"linkedin_live_progress_{run_stamp}_{slug}.json"
    raw_path = logs_dir / f"linkedin_live_raw_inflight_{run_stamp}_{slug}.json"
    return progress_path, raw_path


def _close_page_safely(page: Page | None) -> None:
    if page is None:
        return
    try:
        if not page.is_closed():
            page.close()
    except PlaywrightError:
        pass


def _dir_slug(text: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._ -]+", "", (text or "").strip())
    slug = re.sub(r"\s+", "_", slug).strip("._ ")
    return slug or "item"


def _notes_search(notes: str) -> str:
    match = re.search(r"search=(.*?)(?: window=| insight=|$)", notes)
    return match.group(1).strip() if match else ""


def _notes_window(notes: str) -> str:
    match = re.search(r"window=([^ ]+)", notes)
    return match.group(1).strip() if match else ""


def _cache_key(url_hash_value: str, tc_hash_value: str) -> str:
    return (url_hash_value or "").strip() or (tc_hash_value or "").strip()


def _job_cache_key(job: dict) -> str:
    return _cache_key(str(job.get("url_hash") or ""), str(job.get("tc_hash") or ""))


def _intel_text(notes: str, url: str, fit_score: object | None = None) -> str:
    lines: list[str] = []
    if url:
        lines.append(f"job_link={url}")
    if fit_score not in (None, "", "nan"):
        lines.append(f"fit_score={fit_score}")
    if notes:
        lines.append(str(notes).strip())
    return "\n".join(line for line in lines if line).strip()


def _load_review_cache() -> pd.DataFrame:
    if not JOBS_XLSX.exists():
        return pd.DataFrame(columns=REVIEW_CACHE_COLUMNS)
    try:
        df = pd.read_excel(JOBS_XLSX, sheet_name=REVIEW_CACHE_SHEET_NAME, dtype=str)
    except Exception:
        return pd.DataFrame(columns=REVIEW_CACHE_COLUMNS)
    for col in REVIEW_CACHE_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[REVIEW_CACHE_COLUMNS].fillna("")


def _write_review_cache(rows: list[dict]) -> None:
    if not rows or not JOBS_XLSX.exists():
        return

    from openpyxl import load_workbook

    existing = _load_review_cache()
    merged: dict[str, dict] = {}

    for _, row in existing.iterrows():
        normalized = {col: str(row.get(col, "") or "") for col in REVIEW_CACHE_COLUMNS}
        key = normalized.get("cache_key") or _cache_key(normalized.get("url_hash", ""), normalized.get("tc_hash", ""))
        if key:
            normalized["cache_key"] = key
            merged[key] = normalized

    for row in rows:
        normalized = {col: str(row.get(col, "") or "") for col in REVIEW_CACHE_COLUMNS}
        key = normalized.get("cache_key") or _cache_key(normalized.get("url_hash", ""), normalized.get("tc_hash", ""))
        if key:
            normalized["cache_key"] = key
            merged[key] = normalized

    wb = load_workbook(JOBS_XLSX)
    if REVIEW_CACHE_SHEET_NAME in wb.sheetnames:
        del wb[REVIEW_CACHE_SHEET_NAME]
    ws = wb.create_sheet(REVIEW_CACHE_SHEET_NAME)
    ws.append(REVIEW_CACHE_COLUMNS)
    for row in merged.values():
        ws.append([row.get(col, "") for col in REVIEW_CACHE_COLUMNS])
    wb.save(JOBS_XLSX)


def _terminal_cache_rows(scored_jobs: list[dict]) -> list[dict]:
    rows: list[dict] = []
    reviewed_on = datetime.now().strftime("%Y-%m-%d")
    for job in scored_jobs:
        decision = str(job.get("decision") or "").strip()
        if decision.lower() not in TERMINAL_CACHE_DECISIONS:
            continue
        notes = str(job.get("notes") or "")
        row = {
            "cache_key": _job_cache_key(job),
            "url_hash": str(job.get("url_hash") or ""),
            "tc_hash": str(job.get("tc_hash") or ""),
            "url": str(job.get("url") or ""),
            "company": str(job.get("company") or ""),
            "role_title": str(job.get("role_title") or ""),
            "source": str(job.get("source") or ""),
            "decision": decision,
            "category": str(job.get("category") or ""),
            "fit_score": str(job.get("fit_score") or ""),
            "fit_rationale": str(job.get("fit_rationale") or ""),
            "notes": notes,
            "search_term": _notes_search(notes),
            "time_window": _notes_window(notes),
            "date_reviewed": reviewed_on,
        }
        if row["cache_key"]:
            rows.append(row)
    return rows


def _split_existing_jobs(jobs: list[dict], df_existing: pd.DataFrame) -> tuple[list[dict], list[dict]]:
    existing_url_hashes = {
        str(value).strip()
        for value in df_existing.get("url_hash", pd.Series(dtype=str)).fillna("").tolist()
        if str(value).strip()
    }
    existing_tc_hashes = {
        title_company_hash(str(row.get("role_title") or ""), str(row.get("company") or ""))
        for _, row in df_existing.iterrows()
    }

    unseen: list[dict] = []
    existing_hits: list[dict] = []
    for job in jobs:
        if job["url_hash"] in existing_url_hashes or job["tc_hash"] in existing_tc_hashes:
            existing_hits.append(job)
            continue
        unseen.append(job)
    return unseen, existing_hits


def _split_cached_review_jobs(jobs: list[dict], cache_df: pd.DataFrame) -> tuple[list[dict], list[dict]]:
    cache_by_url: dict[str, dict] = {}
    cache_by_tc: dict[str, dict] = {}

    for _, row in cache_df.iterrows():
        cached = {col: str(row.get(col, "") or "") for col in REVIEW_CACHE_COLUMNS}
        decision = cached.get("decision", "").strip().lower()
        if decision not in TERMINAL_CACHE_DECISIONS:
            continue
        if cached.get("url_hash"):
            cache_by_url[cached["url_hash"]] = cached
        if cached.get("tc_hash"):
            cache_by_tc[cached["tc_hash"]] = cached

    to_score: list[dict] = []
    cache_hits: list[dict] = []
    for job in jobs:
        cached = cache_by_url.get(str(job.get("url_hash") or "")) or cache_by_tc.get(str(job.get("tc_hash") or ""))
        if not cached:
            to_score.append(job)
            continue
        reused = dict(job)
        reused["fit_score"] = cached.get("fit_score") or None
        reused["fit_rationale"] = cached.get("fit_rationale") or None
        reused["decision"] = cached.get("decision") or None
        reused["category"] = cached.get("category") or None
        reused["status"] = "cached_skip"
        reused["notes"] = f"{job.get('notes', '')} cache=review_cache".strip()
        cache_hits.append(reused)
    return to_score, cache_hits


def _export_run_bundle(
    *,
    run_label: str,
    searches: list[tuple[str, str]],
    search_runs: list[dict],
    reviewed_jobs: list[dict],
    fresh_after_dedup: list[dict],
    markdown_report: Path,
    html_report: Path,
    extracted_count: int,
    scored_count: int,
    existing_skip_count: int,
    cache_hits_count: int,
) -> Path:
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    export_dir = RUN_EXPORTS_DIR / f"{stamp}_{_slug(run_label)}"
    accepted_dir = export_dir / "accepted"
    accepted_dir.mkdir(parents=True, exist_ok=True)

    report_md = export_dir / "report.md"
    report_html = export_dir / "report.html"
    shutil.copy2(markdown_report, report_md)
    shutil.copy2(html_report, report_html)

    ranked_jobs = sorted(fresh_after_dedup, key=_job_score_value, reverse=True)
    priority_entries: list[dict] = []

    for rank, job in enumerate(ranked_jobs, start=1):
        company_dir = accepted_dir / _dir_slug(str(job.get("company") or "Unknown"))
        role_dir = company_dir / _dir_slug(str(job.get("role_title") or "Role"))
        role_dir.mkdir(parents=True, exist_ok=True)
        (role_dir / "jd.txt").write_text(str(job.get("jd_text") or "").strip(), encoding="utf-8")
        notes = str(job.get("notes") or "").strip()
        intel_text = _intel_text(notes, str(job.get("url") or ""), job.get("fit_score"))
        if intel_text:
            (role_dir / "intel.txt").write_text(intel_text, encoding="utf-8")
        metadata = {
            "company": job.get("company"),
            "role_title": job.get("role_title"),
            "fit_score": job.get("fit_score"),
            "decision": job.get("decision"),
            "url": job.get("url"),
            "source": job.get("source"),
            "status": job.get("status"),
            "date_found": job.get("date_found"),
            "date_posted": job.get("date_posted"),
            "notes": notes,
            "search_term": _notes_search(notes),
            "time_window": _notes_window(notes),
            "job_id": job.get("id"),
            "priority_rank": rank,
        }
        (role_dir / "metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
        priority_entries.append(
            {
                "priority_rank": rank,
                "company": job.get("company"),
                "role_title": job.get("role_title"),
                "fit_score": job.get("fit_score"),
                "decision": job.get("decision"),
                "url": job.get("url"),
                "bundle_dir": str(role_dir),
            }
        )

    priority_json = export_dir / "priority_order.json"
    priority_txt = export_dir / "priority_order.txt"
    priority_json.write_text(json.dumps(priority_entries, indent=2), encoding="utf-8")
    priority_txt.write_text(
        "\n".join(
            f"{entry['priority_rank']}. {entry['company']} | {entry['role_title']} | score={entry['fit_score']}"
            for entry in priority_entries
        ),
        encoding="utf-8",
    )

    manifest = {
        "run_label": run_label,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "searches": [{"search_term": term, "time_filter": window} for term, window in searches],
        "search_runs": search_runs,
        "counts": {
            "extracted": extracted_count,
            "scored": scored_count,
            "reviewed": len(reviewed_jobs),
            "existing_skipped": existing_skip_count,
            "review_cache_skipped": cache_hits_count,
            "accepted_unique": len(fresh_after_dedup),
        },
        "reports": {
            "markdown": str(report_md),
            "html": str(report_html),
        },
        "priority_files": {
            "json": str(priority_json),
            "text": str(priority_txt),
        },
        "accepted_jobs": [
            {
                "company": job.get("company"),
                "role_title": job.get("role_title"),
                "fit_score": job.get("fit_score"),
                "url": job.get("url"),
                "priority_rank": idx,
                "bundle_dir": str(
                    accepted_dir
                    / _dir_slug(str(job.get("company") or "Unknown"))
                    / _dir_slug(str(job.get("role_title") or "Role"))
                ),
            }
            for idx, job in enumerate(ranked_jobs, start=1)
        ],
    }
    (export_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return export_dir


def _open_linkedin_browser_session(playwright, debug_port: int):
    endpoint = f"http://127.0.0.1:{debug_port}"
    last_error: Exception | None = None
    for attempt in range(2):
        try:
            browser = playwright.chromium.connect_over_cdp(endpoint, timeout=30000)
            if browser.contexts:
                return {
                    "mode": "cdp",
                    "context": browser.contexts[0],
                    "cleanup": lambda: None,
                }
            last_error = RuntimeError("Connected to Chrome, but no browser contexts were available.")
        except PlaywrightError as exc:
            last_error = exc
        if attempt == 0:
            time.sleep(1.0)
    detail = f" Underlying error: {last_error}" if last_error else ""
    raise RuntimeError(
        f"Could not attach to Chrome debug session at {endpoint}. "
        "Launch your normal signed-in Chrome with --remote-debugging-port=9222 and keep it open."
        f"{detail}"
    )


def _slug(text: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9]+", "-", text.strip().lower()).strip("-")
    return value or "run"


def _parse_relative_date(text: str) -> str:
    raw = (text or "").strip().lower()
    now = datetime.now()
    if not raw:
        return now.strftime("%Y-%m-%d")

    patterns: list[tuple[str, timedelta]] = [
        (r"(\d+)\s*hour", timedelta(hours=1)),
        (r"(\d+)\s*day", timedelta(days=1)),
        (r"(\d+)\s*week", timedelta(days=7)),
        (r"(\d+)\s*month", timedelta(days=30)),
    ]
    for pattern, unit in patterns:
        match = re.search(pattern, raw)
        if match:
            count = int(match.group(1))
            return (now - (unit * count)).strftime("%Y-%m-%d")
    if "today" in raw or "just now" in raw:
        return now.strftime("%Y-%m-%d")
    if "yesterday" in raw:
        return (now - timedelta(days=1)).strftime("%Y-%m-%d")
    return now.strftime("%Y-%m-%d")


def _clean_title(text: str) -> str:
    value = re.sub(r"\s+", " ", (text or "")).strip()
    if not value:
        return value

    parts = value.split(" ")
    if len(parts) % 2 == 0:
        half = len(parts) // 2
        if parts[:half] == parts[half:]:
            return " ".join(parts[:half])

    repeated = re.match(r"^(.+?)\s+\1(?:\s+with verification)?$", value, flags=re.I)
    if repeated:
        return repeated.group(1).strip()

    return value


def _scroll_results_list(page: Page) -> int:
    """
    Scroll the jobs results container itself. LinkedIn virtualizes cards inside
    the list, so page-level mouse wheel misses rows on the search-results view.
    Returns the current number of visible job roots after scrolling.
    """
    script = r"""
    () => {
      const selectors = [
        'ul.semantic-search-results-list',
        'ul[class*="semantic-search-results-list"]',
        '.jobs-search-results-list',
        '.scaffold-layout__list ul',
      ];
      const container =
        selectors.map(sel => document.querySelector(sel)).find(Boolean) ||
        document.querySelector('ul:has([data-job-id])');
      if (!container) {
        window.scrollBy(0, 2200);
        return document.querySelectorAll('[data-job-id]').length;
      }
      container.scrollTop = container.scrollHeight;
      return document.querySelectorAll('[data-job-id]').length;
    }
    """
    try:
        visible = page.evaluate(script)
    except PlaywrightError:
        page.mouse.wheel(0, 2200)
        visible = page.locator("[data-job-id]").count()
    _human_pause(page, 700, 1200)
    return int(visible or 0)


def _extract_about_job_text(page: Page) -> str:
    def _extract_text_via_js() -> list[str]:
        try:
            payload = page.evaluate(
                """
                () => {
                  const texts = [];
                  const seen = new Set();
                  const pushNode = (node) => {
                    if (!node) return;
                    const text = (node.innerText || node.textContent || '').trim();
                    if (!text || seen.has(text)) return;
                    seen.add(text);
                    texts.push(text);
                  };
                  pushNode(document.querySelector('main'));
                  pushNode(document.querySelector('[role="main"]'));
                  pushNode(document.body);
                  return texts;
                }
                """
            )
            return [str(item or "").strip() for item in (payload or []) if str(item or "").strip()]
        except PlaywrightError:
            return []

    def _slice_job_text(raw_text: str) -> str:
        normalized = re.sub(r"\n{3,}", "\n\n", (raw_text or "").strip())
        if not normalized:
            return ""

        start_markers = [
            "About the job",
            "Job Summary",
            "Role Overview",
            "About this role",
            "Position Summary",
            "Overview",
            "What you'll do",
            "What you will do",
            "Primary Duties And Responsibilities",
            "Responsibilities",
        ]
        end_markers = [
            "Seniority level",
            "Employment type",
            "Job function",
            "Industries",
            "Referrals increase your chances",
            "Set alert for similar jobs",
            "People you can reach out to",
            "Meet the hiring team",
            "About the company",
            "Get notified about new",
            "Show less",
        ]

        start_idx = None
        matched_marker = ""
        for marker in start_markers:
            idx = normalized.lower().find(marker.lower())
            if idx >= 0 and (start_idx is None or idx < start_idx):
                start_idx = idx
                matched_marker = marker

        if start_idx is None:
            return normalized

        start = start_idx + len(matched_marker)
        body = normalized[start:].strip()
        end_positions = [
            body.lower().find(marker.lower())
            for marker in end_markers
            if body.lower().find(marker.lower()) > 0
        ]
        if end_positions:
            body = body[: min(end_positions)].strip()
        return body

    for raw_text in _extract_text_via_js():
        sliced = _slice_job_text(raw_text)
        if sliced:
            return sliced

    candidates = [
        page.locator("main").first,
        page.locator("body").first,
    ]
    for locator in candidates:
        try:
            if locator.count() == 0:
                continue
            text = locator.inner_text(timeout=3000).strip()
            sliced = _slice_job_text(text)
            if sliced:
                return sliced
        except PlaywrightError:
            continue
    return ""


def _normalize_jd_text(text: str) -> str:
    return re.sub(r"\n{3,}", "\n\n", (text or "").strip())


def _jd_quality_issue(text: str) -> str:
    normalized = _normalize_jd_text(text)
    if not normalized:
        return "empty"

    lower = normalized.lower()
    chrome_hits = sum(marker in lower for marker in JD_CHROME_MARKERS)
    section_hits = sum(marker in lower for marker in JD_SECTION_MARKERS)

    if normalized.startswith("Looking for talent?"):
        return "linkedin_chrome"
    if "linkedin corporation ©" in lower and section_hits == 0:
        return "linkedin_chrome"
    if chrome_hits >= 3 and section_hits == 0:
        return "linkedin_chrome"
    if chrome_hits >= 2 and len(normalized) < 2200 and section_hits == 0:
        return "linkedin_chrome"
    return ""


def _expand_job_description(page: Page) -> None:
    expanders = [
        page.get_by_role("button", name=re.compile("show more", re.I)),
        page.locator(".show-more-less-html__button"),
        page.locator("button[aria-label*='show more' i]"),
    ]
    for locator in expanders:
        try:
            count = min(locator.count(), 3)
        except PlaywrightError:
            continue
        for idx in range(count):
            try:
                locator.nth(idx).click(timeout=1000)
                page.wait_for_timeout(400)
            except PlaywrightError:
                continue


def _wait_for_job_description(page: Page, timeout_ms: int = 3500) -> None:
    selectors = [
        ".jobs-search__job-details--wrapper",
        ".jobs-box__html-content",
        "#job-details",
        ".jobs-description",
        ".jobs-description-content__text",
        ".show-more-less-html__markup",
    ]
    deadline = time.monotonic() + (timeout_ms / 1000)
    while time.monotonic() < deadline:
        for selector in selectors:
            try:
                locator = page.locator(selector).first
                if locator.count() and locator.inner_text(timeout=400).strip():
                    return
            except PlaywrightError:
                continue
        page.wait_for_timeout(350)


def _safe_goto(page: Page, url: str, timeout_ms: int = 30000) -> bool:
    def _looks_loaded() -> bool:
        try:
            current_url = page.url.lower()
            if "linkedin.com/authwall" in current_url or "linkedin.com/login" in current_url:
                return False
            if "linkedin.com/jobs" not in current_url:
                return False
            page.wait_for_timeout(1200)
            if page.locator("[data-job-id]").count() > 0:
                return True
            body_text = page.locator("body").inner_text(timeout=1500)
            body_text = body_text.lower()
            if "join linkedin" in body_text or "agree & join" in body_text or "already on linkedin? sign in" in body_text:
                return False
            return "results" in body_text or "jobs" in body_text
        except PlaywrightError:
            return False

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
        return _looks_loaded()
    except PlaywrightTimeoutError:
        try:
            page.goto(url, wait_until="commit", timeout=min(timeout_ms, 12000))
            return _looks_loaded()
        except PlaywrightError:
            return False
    except PlaywrightError:
        return False


def _body_preview(page: Page) -> str:
    try:
        text = page.locator("body").inner_text(timeout=2000)
    except PlaywrightError:
        return ""
    return " ".join(text.split())[:400]


def _is_authwall_or_login(page: Page) -> bool:
    current_url = page.url.lower()
    if "linkedin.com/authwall" in current_url or "linkedin.com/login" in current_url:
        return True
    preview = _body_preview(page).lower()
    return any(
        token in preview
        for token in (
            "join linkedin",
            "sign in",
            "agree & join",
            "new to linkedin",
            "already on linkedin?",
        )
    )


def _semantic_fuzzy_count(page: Page) -> int | None:
    code_nodes = page.locator("code")
    total = min(code_nodes.count(), 120)
    for idx in range(total):
        try:
            text = code_nodes.nth(idx).inner_text(timeout=500)
        except PlaywrightError:
            continue
        if "jobsDashJobCardsBySemanticSearch" not in text or "fuzzyCount" not in text:
            continue
        match = re.search(r'"fuzzyCount":\{.*?"text":"(\d+)\s+results"', text)
        if match:
            return int(match.group(1))
    return None


def _heading_results_count(page: Page) -> int | None:
    selectors = [
        "h1",
        ".jobs-search-results-list__title-heading",
        ".jobs-search-box__results-header",
        "[data-view-name='search-results-page-title']",
    ]
    pattern = re.compile(r"\b(\d[\d,]*)\+?\s+results?\b", re.I)
    for selector in selectors:
        locator = page.locator(selector)
        total = min(locator.count(), 5)
        for idx in range(total):
            try:
                text = locator.nth(idx).inner_text(timeout=500)
            except PlaywrightError:
                continue
            match = pattern.search(text or "")
            if match:
                return int(match.group(1).replace(",", ""))
    return None


def _ui_observed_results_count(page: Page) -> int | None:
    return _semantic_fuzzy_count(page) or _heading_results_count(page)


def _session_preflight(context, target_url: str = "https://www.linkedin.com/feed/") -> dict:
    page_count_before = len(context.pages)
    page = context.new_page()
    page.set_default_timeout(15000)
    try:
        page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(1500)
        cookies = context.cookies(["https://www.linkedin.com"])
        has_li_at = any(cookie.get("name") == "li_at" for cookie in cookies)
        logged_in = _looks_logged_in(page)
        authwall = _is_authwall_or_login(page)
        return {
            "ok": logged_in and not authwall,
            "target_url": target_url,
            "current_url": page.url,
            "title": page.title(),
            "logged_in_heuristic": logged_in,
            "authwall_or_login": authwall,
            "has_li_at_cookie": has_li_at,
            "cookie_names": sorted(cookie.get("name", "") for cookie in cookies),
            "body_preview": _body_preview(page),
            "context_pages_before": page_count_before,
        }
    finally:
        _close_page_safely(page)


def _job_score_value(job: dict) -> float:
    try:
        return float(job.get("fit_score"))
    except (TypeError, ValueError):
        return -1.0


def _report_gate_label(job: dict, accepted_urls: set[str]) -> str:
    gate = str(job.get("__report_gate") or "").strip()
    if gate:
        return gate
    return "accepted" if job.get("url") in accepted_urls else "dropped"


def _report_reason_tag(job: dict) -> str:
    gate = str(job.get("__report_gate") or "").strip().lower()
    if gate == "existing":
        return "existing"

    decision = str(job.get("decision") or "").strip().lower()
    if decision == "proceed":
        return ""
    rationale = str(job.get("fit_rationale") or "").lower()
    role_title = str(job.get("role_title") or "").lower()

    if decision == "error":
        if "no jd text" in rationale:
            return "jd-missing"
        if "invalid jd capture" in rationale or "linkedin shell" in rationale:
            return "bad-jd"
        return "extract-fail"

    checks = [
        ("undergrad-only", ("undergraduate students", "working towards a bachelor's", "currently enrolled undergraduate", "rising senior", "incoming junior", "incoming senior")),
        ("language", ("japanese", "bilingual", "full professional proficiency", "native or full professional proficiency")),
        ("location", ("full on-site availability", "onsite in", "on-site in", "location mismatch", "must be available to work onsite")),
        ("cpt-explicit", ("no cpt", "no opt", "f-1", "f1", "permanent basis", "permanent work authorization", "us citizen", "green card", "security clearance", "us person")),
        ("visa-ambig", ("no sponsorship", "visa sponsorship", "without visa sponsorship", "authorized to work without sponsorship", "no employment sponsorship")),
        ("senior-fulltime", ("4+ years required", "3+ years required", "8 years required", "full-time hire", "senior hire", "not an internship")),
        ("level", ("level mismatch", "2–4 years", "2-4 years")),
        ("marketing", ("marketing intern", "product marketing")),
        ("ops-heavy", ("ops-heavy", "operations support", "process management", "execution-focused")),
        ("role-mismatch", ("role type mismatch", "role-type mismatch", "hard mismatch")),
        ("non-tech", ("non-tech", "traditional non-tech", "retail", "timeshare", "hospitality")),
        ("no-pm", ("no product ownership", "minimal product ownership", "no strategic decision-making")),
        ("corp-dev", ("corporate development", "business development", "sales operations")),
        ("thin-jd", ("under-specified", "insufficient information", "cannot proceed without substantive role details")),
    ]
    for tag, needles in checks:
        if any(needle in rationale for needle in needles):
            return tag

    if "marketing" in role_title:
        return "marketing"
    if decision == "deprioritize":
        return "weak-fit"
    if decision == "reject":
        return "reject"
    return ""


def _render_report_markdown(
    *,
    run_label: str,
    searches: list[tuple[str, str]],
    search_runs: list[dict],
    scored_jobs: list[dict],
    all_reviewed_jobs: list[dict],
    accepted_for_write: list[dict],
    fresh_after_dedup: list[dict],
    extracted_count: int,
    scored_count: int,
    existing_skip_count: int,
    cache_hits_count: int,
) -> str:
    decisions: dict[str, int] = {}
    for job in scored_jobs:
        key = str(job.get("decision") or "Unknown")
        decisions[key] = decisions.get(key, 0) + 1

    top_jobs = sorted(
        [job for job in scored_jobs if _job_score_value(job) >= 0],
        key=_job_score_value,
        reverse=True,
    )[:10]

    lines = [
        f"# LinkedIn Discovery Batch Report: {run_label}",
        "",
        "## Summary",
        f"- Searches: {', '.join(f'{term} ({TIME_LABELS.get(window, window)})' for term, window in searches)}",
        f"- Jobs extracted: {extracted_count}",
        f"- Jobs scored this run: {scored_count}",
        f"- Jobs skipped as existing rows: {existing_skip_count}",
        f"- Jobs skipped from review cache: {cache_hits_count}",
        f"- Jobs accepted for write gate: {len(accepted_for_write)}",
        f"- Jobs written after dedup: {len(fresh_after_dedup)}",
        f"- Decisions: {', '.join(f'{k}={v}' for k, v in sorted(decisions.items())) or 'none'}",
        "",
        "## Search Counts",
    ]
    if search_runs:
        for run in search_runs:
            ui_count = run.get("ui_observed_count")
            mismatch = run.get("count_mismatch")
            mismatch_note = " mismatch=yes" if mismatch else ""
            lines.append(
                f"- {run.get('search_term')} ({TIME_LABELS.get(run.get('time_filter'), run.get('time_filter'))}): "
                f"ui_reported_total={ui_count if ui_count is not None else 'unknown'}, "
                f"extracted_unique={run.get('extracted_count')}, "
                f"route={run.get('route_used')}{mismatch_note}"
            )
    else:
        lines.append("- No per-search telemetry recorded.")

    lines.extend([
        "",
        "## Top Rated",
    ])

    if top_jobs:
        lines.extend([
            "| Score | Decision | Company | Role | Search | Window |",
            "|---|---|---|---|---|---|",
        ])
        for job in top_jobs:
            lines.append(
                f"| {job.get('fit_score', '')} | {job.get('decision', '')} | "
                f"{job.get('company', '')} | {job.get('role_title', '')} | "
                f"{re.search(r'search=([^ ]+.*?)(?: window=| insight=|$)', str(job.get('notes', ''))).group(1) if re.search(r'search=([^ ]+.*?)(?: window=| insight=|$)', str(job.get('notes', ''))) else ''} | "
                f"{re.search(r'window=([^ ]+)', str(job.get('notes', ''))).group(1) if re.search(r'window=([^ ]+)', str(job.get('notes', ''))) else ''} |"
            )
    else:
        lines.append("No scored jobs available.")

    lines.extend(["", "## All Reviewed Jobs", "| Score | Decision | Reason | Write Gate | Company | Role | Source Search |", "|---|---|---|---|---|---|---|"])
    accepted_urls = {job.get("url") for job in accepted_for_write}
    for job in sorted(all_reviewed_jobs, key=_job_score_value, reverse=True):
        notes = str(job.get("notes", ""))
        search_match = re.search(r"search=(.*?)(?: window=| insight=|$)", notes)
        source_search = search_match.group(1) if search_match else ""
        lines.append(
            f"| {job.get('fit_score', '') if job.get('fit_score') is not None else ''} | "
            f"{job.get('decision', '')} | "
            f"{_report_reason_tag(job)} | "
            f"{_report_gate_label(job, accepted_urls)} | "
            f"{job.get('company', '')} | {job.get('role_title', '')} | {source_search} |"
        )
    lines.append("")
    return "\n".join(lines)


def _render_report_html(
    *,
    run_label: str,
    searches: list[tuple[str, str]],
    search_runs: list[dict],
    scored_jobs: list[dict],
    all_reviewed_jobs: list[dict],
    accepted_for_write: list[dict],
    fresh_after_dedup: list[dict],
    extracted_count: int,
    scored_count: int,
    existing_skip_count: int,
    cache_hits_count: int,
) -> str:
    decisions: dict[str, int] = {}
    for job in scored_jobs:
        key = str(job.get("decision") or "Unknown")
        decisions[key] = decisions.get(key, 0) + 1

    top_jobs = sorted(
        [job for job in scored_jobs if _job_score_value(job) >= 0],
        key=_job_score_value,
        reverse=True,
    )[:10]
    accepted_urls = {job.get("url") for job in accepted_for_write}

    def search_of(job: dict) -> str:
        notes = str(job.get("notes", ""))
        match = re.search(r"search=(.*?)(?: window=| insight=|$)", notes)
        return match.group(1) if match else ""

    def rows_html(rows: list[dict], include_gate: bool) -> str:
        out = []
        for job in rows:
            cells = [
                html.escape("" if job.get("fit_score") is None else str(job.get("fit_score"))),
                html.escape(str(job.get("decision", ""))),
                html.escape(_report_reason_tag(job)),
            ]
            if include_gate:
                cells.append(_report_gate_label(job, accepted_urls))
            cells.extend([
                html.escape(str(job.get("company", ""))),
                html.escape(str(job.get("role_title", ""))),
                html.escape(search_of(job)),
            ])
            out.append("<tr>" + "".join(f"<td>{cell}</td>" for cell in cells) + "</tr>")
        return "\n".join(out)

    summary_items = [
        f"Searches: {', '.join(f'{term} ({TIME_LABELS.get(window, window)})' for term, window in searches)}",
        f"Jobs extracted: {extracted_count}",
        f"Jobs scored this run: {scored_count}",
        f"Jobs skipped as existing rows: {existing_skip_count}",
        f"Jobs skipped from review cache: {cache_hits_count}",
        f"Jobs accepted for write gate: {len(accepted_for_write)}",
        f"Jobs written after dedup: {len(fresh_after_dedup)}",
        f"Decisions: {', '.join(f'{k}={v}' for k, v in sorted(decisions.items())) or 'none'}",
    ]
    search_count_items = [
        (
            f"{run.get('search_term')} ({TIME_LABELS.get(run.get('time_filter'), run.get('time_filter'))}): "
            f"ui_reported_total={run.get('ui_observed_count') if run.get('ui_observed_count') is not None else 'unknown'}, "
            f"extracted_unique={run.get('extracted_count')}, route={run.get('route_used')}"
            + (" mismatch=yes" if run.get("count_mismatch") else "")
        )
        for run in search_runs
    ]

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>{html.escape(f"LinkedIn Discovery Batch Report: {run_label}")}</title>
  <style>
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      max-width: 980px;
      margin: 32px auto;
      padding: 0 20px 40px;
      color: #1f2937;
      background: #f8fafc;
      line-height: 1.5;
    }}
    h1, h2 {{ color: #0f172a; }}
    .card {{ background: white; padding: 16px 18px; border-radius: 14px; margin-bottom: 18px; box-shadow: 0 1px 2px rgba(15, 23, 42, 0.08); }}
    ul {{ margin: 0; padding-left: 20px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
    th, td {{ padding: 10px 8px; border-bottom: 1px solid #e2e8f0; text-align: left; vertical-align: top; }}
    th {{ background: #eff6ff; color: #1d4ed8; position: sticky; top: 0; }}
    .muted {{ color: #64748b; }}
  </style>
</head>
<body>
  <h1>LinkedIn Discovery Batch Report: {html.escape(run_label)}</h1>
  <div class="card">
    <h2>Summary</h2>
    <ul>
      {''.join(f'<li>{html.escape(item)}</li>' for item in summary_items)}
    </ul>
  </div>
  <div class="card">
    <h2>Search Counts</h2>
    <ul>
      {''.join(f'<li>{html.escape(item)}</li>' for item in search_count_items) if search_count_items else '<li class="muted">No per-search telemetry recorded.</li>'}
    </ul>
  </div>
  <div class="card">
    <h2>Top Rated</h2>
    <table>
      <thead>
        <tr><th>Score</th><th>Decision</th><th>Company</th><th>Role</th><th>Search</th></tr>
      </thead>
      <tbody>
        {rows_html(top_jobs, include_gate=False) if top_jobs else '<tr><td colspan="5" class="muted">No scored jobs available.</td></tr>'}
      </tbody>
    </table>
  </div>
  <div class="card">
    <h2>All Reviewed Jobs</h2>
    <table>
      <thead>
        <tr><th>Score</th><th>Decision</th><th>Reason</th><th>Write Gate</th><th>Company</th><th>Role</th><th>Search</th></tr>
      </thead>
      <tbody>
        {rows_html(sorted(all_reviewed_jobs, key=_job_score_value, reverse=True), include_gate=True)}
      </tbody>
    </table>
  </div>
</body>
</html>
"""


def _write_batch_report(
    *,
    run_label: str,
    searches: list[tuple[str, str]],
    search_runs: list[dict],
    scored_jobs: list[dict],
    all_reviewed_jobs: list[dict],
    accepted_for_write: list[dict],
    fresh_after_dedup: list[dict],
    extracted_count: int,
    scored_count: int,
    existing_skip_count: int,
    cache_hits_count: int,
) -> tuple[Path, Path]:
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    base = f"linkedin_live_report_{stamp}_{_slug(run_label)}"
    md_path = _ensure_logs_dir() / f"{base}.md"
    html_path = _ensure_logs_dir() / f"{base}.html"
    md = _render_report_markdown(
        run_label=run_label,
        searches=searches,
        search_runs=search_runs,
        scored_jobs=scored_jobs,
        all_reviewed_jobs=all_reviewed_jobs,
        accepted_for_write=accepted_for_write,
        fresh_after_dedup=fresh_after_dedup,
        extracted_count=extracted_count,
        scored_count=scored_count,
        existing_skip_count=existing_skip_count,
        cache_hits_count=cache_hits_count,
    )
    md_path.write_text(md, encoding="utf-8")
    html_path.write_text(
        _render_report_html(
            run_label=run_label,
            searches=searches,
            search_runs=search_runs,
            scored_jobs=scored_jobs,
            all_reviewed_jobs=all_reviewed_jobs,
            accepted_for_write=accepted_for_write,
            fresh_after_dedup=fresh_after_dedup,
            extracted_count=extracted_count,
            scored_count=scored_count,
            existing_skip_count=existing_skip_count,
            cache_hits_count=cache_hits_count,
        ),
        encoding="utf-8",
    )
    return md_path, html_path


def _looks_logged_in(page: Page) -> bool:
    if page.is_closed():
        return False
    try:
        current_url = page.url.lower()
        if "linkedin.com/authwall" in current_url or "linkedin.com/login" in current_url:
            return False
        checks = [
            ("feed", "linkedin.com/feed" in current_url),
            ("jobs", "linkedin.com/jobs" in current_url),
            ("me icon", page.locator('[data-control-name="nav.settings"]').count() > 0),
            ("global nav", page.locator("nav.global-nav").count() > 0),
        ]
    except PlaywrightError:
        return False
    return any(ok for _, ok in checks)


def _extract_visible_cards(page: Page) -> list[dict]:
    script = """
    () => {
      const normalize = (value) => value ? value.replace(/\\s+/g, ' ').trim() : '';
      const cleanRepeated = (value) => {
        const normalized = normalize(value);
        if (!normalized || normalized.length % 2 !== 0) return normalized;
        const half = normalized.length / 2;
        const first = normalized.slice(0, half);
        const second = normalized.slice(half);
        return first === second ? first : normalized;
      };
      const seen = new Set();
      const cards = [];
      const cardNodes = Array.from(document.querySelectorAll('[data-job-id]'));
      for (const node of cardNodes) {
        const jobId = node.getAttribute('data-job-id');
        const url = jobId ? `https://www.linkedin.com/jobs/view/${jobId}/` : '';
        if (!url || seen.has(url)) continue;
        seen.add(url);

        const titleNode =
          node.querySelector('.job-card-job-posting-card-wrapper__title strong') ||
          node.querySelector('.job-card-job-posting-card-wrapper__title') ||
          node.querySelector('.job-card-list__title') ||
          node.querySelector('.job-card-container__link');
        const companyNode =
          node.querySelector('.artdeco-entity-lockup__subtitle') ||
          node.querySelector('.job-card-container__company-name') ||
          node.querySelector('.artdeco-entity-lockup__primary-subtitle');
        const metaTexts = Array.from(node.querySelectorAll(
          '.job-card-container__metadata-item, .artdeco-entity-lockup__caption, .job-card-container__footer-item, .artdeco-entity-lockup__metadata, time'
        )).map(el => normalize(el.textContent)).filter(Boolean);
        const location = metaTexts.find(text =>
          text &&
          !/(ago|applicant|applicants|clicked apply|viewed|easy apply|promoted|response|benefit|school alum|company alumni|connections? work here)/i.test(text)
        ) || '';
        const listedAt = metaTexts.find(text =>
          /(ago|today|yesterday|viewed|reposted|within the past)/i.test(text)
        ) || '';

        cards.push({
          title: cleanRepeated(titleNode ? titleNode.textContent : ''),
          company: normalize(companyNode ? companyNode.textContent : ''),
          location,
          listed_at: listedAt,
          url,
        });
      }
      return cards;
    }
    """
    data = page.evaluate(script)
    return [row for row in data if row.get("title") and row.get("company") and row.get("url")]


def _open_job_details(page: Page, job_url: str, detail_page: Page | None = None) -> tuple[str, str]:
    owns_page = detail_page is None
    detail_page = detail_page or page.context.new_page()
    detail_page.set_default_timeout(5000)
    try:
        last_jd_text = ""
        last_insight_text = ""
        for attempt in range(1, 4):
            detail_page.goto(job_url, wait_until="domcontentloaded", timeout=8000)
            detail_page.wait_for_timeout(700 + (attempt * 300))
            _wait_for_job_description(detail_page, timeout_ms=2500 + (attempt * 1000))
            _expand_job_description(detail_page)

            extracted = detail_page.evaluate(
                """
                () => {
                  const pickText = (selectors) => {
                    for (const selector of selectors) {
                      const node = document.querySelector(selector);
                      if (!node) continue;
                      const text = (node.innerText || '').replace(/\\s+/g, ' ').trim();
                      if (text) return text;
                    }
                    return '';
                  };

                  let jdText = pickText([
                    '.jobs-search__job-details--wrapper',
                    '.jobs-box__html-content',
                    '#job-details',
                    '.jobs-description',
                    '.jobs-description-content__text',
                    '.show-more-less-html__markup',
                  ]);

                  let insightText = pickText([
                    '.jobs-unified-top-card__job-insight',
                    '.job-details-jobs-unified-top-card__primary-description-container',
                    '.jobs-unified-top-card__primary-description-container',
                    '.jobs-unified-top-card__subtitle-primary-grouping',
                  ]);

                  if (!insightText) {
                    const main = document.querySelector('main');
                    const mainText = (main?.innerText || '').replace(/\\s+/g, ' ').trim();
                    if (mainText) {
                      const aboutIdx = mainText.indexOf('About the job');
                      insightText = aboutIdx > 0 ? mainText.slice(0, aboutIdx).trim() : mainText.slice(0, 400).trim();
                    }
                  }

                  return { jdText, insightText };
                }
                """
            )

            jd_text = (extracted or {}).get("jdText", "").strip()
            if not jd_text:
                jd_text = _extract_about_job_text(detail_page)
            jd_text = _normalize_jd_text(jd_text)
            insight_text = (extracted or {}).get("insightText", "").strip()

            last_jd_text = jd_text
            last_insight_text = insight_text
            issue = _jd_quality_issue(jd_text)
            if not issue:
                return jd_text, insight_text

            if issue == "linkedin_chrome":
                detail_page.wait_for_timeout(800 * attempt)
                continue

        return "", last_insight_text
    except PlaywrightError:
        return "", ""
    finally:
        if owns_page:
            detail_page.close()


def _repair_scraped_cards(
    *,
    page: Page,
    detail_page: Page,
    cards: list[LinkedInJobCard],
    quiet: bool,
    progress_callback: Callable[[dict], None] | None = None,
    max_jobs: int = JD_REPAIR_MAX_JOBS,
) -> JdRepairSummary:
    candidates = [card for card in cards if _jd_quality_issue(card.jd_text)]
    if not candidates:
        return JdRepairSummary(candidates=0, attempted=0, repaired=0, remaining_failed=0)

    attempted = 0
    repaired = 0
    repair_batch = candidates[:max(max_jobs, 0)]

    if not quiet:
        print(
            f"\n[repair] Retrying JD extraction for {len(repair_batch)} of "
            f"{len(candidates)} candidate jobs before scoring..."
        )

    for idx, card in enumerate(repair_batch, start=1):
        attempted += 1
        if progress_callback is not None:
            progress_callback(
                {
                    "event": "jd_repair_attempt",
                    "repair_index": idx,
                    "repair_total": len(repair_batch),
                    "repair_url": card.url,
                    "repair_title": card.title,
                    "repair_company": card.company,
                }
            )
        try:
            jd_text, insight_text = _open_job_details(page, card.url, detail_page=detail_page)
        except PlaywrightError:
            jd_text, insight_text = "", ""

        if _jd_quality_issue(jd_text):
            continue

        card.jd_text = jd_text
        if insight_text.strip():
            card.insight = insight_text.strip()
        repaired += 1
        if progress_callback is not None:
            progress_callback(
                {
                    "event": "jd_repair_success",
                    "repair_index": idx,
                    "repair_total": len(repair_batch),
                    "repair_url": card.url,
                    "repair_title": card.title,
                    "repair_company": card.company,
                }
            )

    remaining_failed = sum(1 for card in cards if _jd_quality_issue(card.jd_text))
    if not quiet:
        print(
            f"[repair] JD repair complete: repaired {repaired}/{attempted} attempted "
            f"({remaining_failed} still missing/invalid)"
        )
    return JdRepairSummary(
        candidates=len(candidates),
        attempted=attempted,
        repaired=repaired,
        remaining_failed=remaining_failed,
    )


def _goto_next_page(page: Page) -> bool:
    candidates = [
        page.get_by_role("button", name=re.compile("next", re.I)),
        page.locator("button[aria-label*='next' i]"),
        page.locator("button.jobs-search-pagination__indicator-button--next"),
        page.locator(".artdeco-pagination__button--next"),
    ]
    for locator in candidates:
        try:
            if locator.count() == 0:
                continue
            button = locator.first
            if button.is_disabled():
                return False
            button.click(timeout=5000)
            _human_pause(page, 1400, 2400)
            return True
        except PlaywrightError:
            try:
                handle = button.element_handle(timeout=2000)
                if handle is None:
                    continue
                page.evaluate("(el) => el.click()", handle)
                _human_pause(page, 1400, 2400)
                return True
            except PlaywrightError:
                continue
    return False


def scrape_search(
    page: Page,
    search_term: str,
    time_filter: str,
    limit_per_search: int | None,
    pages: int | None,
    count_only: bool = False,
    detail_page: Page | None = None,
    progress_callback: Callable[[dict], None] | None = None,
) -> SearchRunResult:
    def _run_single_url(url: str) -> tuple[list[LinkedInJobCard], int | None]:
        cards: list[LinkedInJobCard] = []
        seen_urls: set[str] = set()
        observed_ui_count: int | None = None

        use_offset_paging = "linkedin.com/jobs/search-results/" in url
        route_used = "search-results" if use_offset_paging else "jobs-search-fallback"
        if use_offset_paging:
            if not _safe_goto(page, url):
                return cards, observed_ui_count
            _human_pause(page, 1600, 2400)
            if not _looks_logged_in(page):
                raise RuntimeError(
                    "LinkedIn does not appear logged in in the connected Chrome session. "
                    "Open your logged-in LinkedIn jobs page in the debug-enabled Chrome window first."
                )
            page.wait_for_timeout(1500)
            page_signatures: set[tuple[str, ...]] = set()
            start = 0
            page_index = 0
            stagnant_pages = 0
            while True:
                if pages is not None and page_index >= max(pages, 1):
                    break
                target_url = _with_start_param(url, start) if start else url
                if not _safe_goto(page, target_url):
                    break
                _human_pause(page, 1600, 2400)

                if not _looks_logged_in(page):
                    raise RuntimeError(
                        "LinkedIn does not appear logged in in the connected Chrome session. "
                        "Open your logged-in LinkedIn jobs page in the debug-enabled Chrome window first."
                    )

                page.wait_for_timeout(2000)
                page_ui_count = _ui_observed_results_count(page)
                if page_ui_count is not None:
                    observed_ui_count = max(observed_ui_count or 0, page_ui_count)
                if progress_callback is not None:
                    progress_callback(
                        {
                            "event": "page_loaded",
                            "search_term": search_term,
                            "time_filter": time_filter,
                            "route": route_used,
                            "page_index": page_index,
                            "start": start,
                            "search_extracted": len(cards),
                            "ui_observed_count": observed_ui_count,
                            "cards_snapshot": cards,
                        }
                    )
                initial_visible = _extract_visible_cards(page)
                if not initial_visible:
                    break
                signature = tuple(card["url"] for card in initial_visible[:10])
                if signature in page_signatures:
                    break
                page_signatures.add(signature)

                before_page_count = len(seen_urls)
                page_seen_urls: set[str] = set()
                stagnant_scrolls = 0
                for _ in range(20):
                    before_count = len(seen_urls)
                    visible = _extract_visible_cards(page)
                    for summary in visible:
                        job_url = summary["url"]
                        page_seen_urls.add(job_url)
                        if job_url in seen_urls:
                            continue
                        if count_only:
                            jd_text, insight_text = "", ""
                        else:
                            try:
                                jd_text, insight_text = _open_job_details(page, job_url, detail_page=detail_page)
                            except PlaywrightError:
                                jd_text, insight_text = "", ""

                        seen_urls.add(job_url)
                        cards.append(
                            LinkedInJobCard(
                                search_term=search_term,
                                time_filter=time_filter,
                                title=_clean_title(summary["title"]),
                                company=summary["company"],
                                location=summary["location"],
                                url=job_url,
                                listed_at=summary["listed_at"],
                                insight=insight_text,
                                jd_text=jd_text,
                            )
                        )
                        if limit_per_search is not None and len(cards) >= limit_per_search:
                            return cards, observed_ui_count

                    if progress_callback is not None and len(seen_urls) > before_count:
                        progress_callback(
                            {
                                "event": "new_cards",
                                "search_term": search_term,
                                "time_filter": time_filter,
                                "route": route_used,
                                "page_index": page_index,
                                "start": start,
                                "search_extracted": len(cards),
                                "ui_observed_count": observed_ui_count,
                                "cards_snapshot": cards,
                            }
                        )
                    if len(seen_urls) == before_count:
                        stagnant_scrolls += 1
                    else:
                        stagnant_scrolls = 0
                    if stagnant_scrolls >= 2:
                        break
                    _scroll_results_list(page)

                page_growth = len(seen_urls) - before_page_count
                if page_growth <= 0:
                    stagnant_pages += 1
                else:
                    stagnant_pages = 0

                if stagnant_pages >= SEARCH_RESULTS_MAX_STAGNANT_OFFSET_PAGES:
                    break

                if observed_ui_count is not None and len(seen_urls) >= observed_ui_count:
                    break

                page_size = max(
                    len(page_seen_urls),
                    len(initial_visible),
                    SEARCH_RESULTS_OFFSET_PAGE_SIZE,
                )
                start += page_size
                page_index += 1
            return cards, observed_ui_count

        page_index = 0
        while True:
            if not _safe_goto(page, url):
                break
            _human_pause(page, 1600, 2400)

            if not _looks_logged_in(page):
                raise RuntimeError(
                    "LinkedIn does not appear logged in in the connected Chrome session. "
                    "Open your logged-in LinkedIn jobs page in the debug-enabled Chrome window first."
                )

            page.wait_for_timeout(2000)
            page_ui_count = _ui_observed_results_count(page)
            if page_ui_count is not None:
                observed_ui_count = max(observed_ui_count or 0, page_ui_count)
            if progress_callback is not None:
                progress_callback(
                    {
                        "event": "page_loaded",
                        "search_term": search_term,
                        "time_filter": time_filter,
                        "route": route_used,
                        "page_index": page_index,
                        "search_extracted": len(cards),
                        "ui_observed_count": observed_ui_count,
                        "cards_snapshot": cards,
                    }
                )
            stagnant_scrolls = 0
            for _ in range(20):
                before_count = len(seen_urls)
                visible = _extract_visible_cards(page)
                for summary in visible:
                    job_url = summary["url"]
                    if job_url in seen_urls:
                        continue
                    if count_only:
                        jd_text, insight_text = "", ""
                    else:
                        try:
                            jd_text, insight_text = _open_job_details(page, job_url, detail_page=detail_page)
                        except PlaywrightError:
                            jd_text, insight_text = "", ""

                    seen_urls.add(job_url)
                    cards.append(
                        LinkedInJobCard(
                            search_term=search_term,
                            time_filter=time_filter,
                            title=_clean_title(summary["title"]),
                            company=summary["company"],
                            location=summary["location"],
                            url=job_url,
                            listed_at=summary["listed_at"],
                            insight=insight_text,
                            jd_text=jd_text,
                        )
                    )
                    if limit_per_search is not None and len(cards) >= limit_per_search:
                        return cards, observed_ui_count

                if progress_callback is not None and len(seen_urls) > before_count:
                    progress_callback(
                        {
                            "event": "new_cards",
                            "search_term": search_term,
                            "time_filter": time_filter,
                            "route": route_used,
                            "page_index": page_index,
                            "search_extracted": len(cards),
                            "ui_observed_count": observed_ui_count,
                            "cards_snapshot": cards,
                        }
                    )
                if len(seen_urls) == before_count:
                    stagnant_scrolls += 1
                else:
                    stagnant_scrolls = 0
                if stagnant_scrolls >= 2:
                    break
                _scroll_results_list(page)

            if pages is not None and page_index + 1 >= max(pages, 1):
                break
            if limit_per_search is not None and len(cards) >= limit_per_search:
                break
            moved = _goto_next_page(page)
            if not moved:
                break
            page_index += 1
        return cards, observed_ui_count

    primary_url = _jobs_search_url(search_term, time_filter)
    fallback_url = _jobs_search_fallback_url(search_term, time_filter)
    primary_cards, primary_ui_count = _run_single_url(primary_url)
    if not _should_try_fallback(len(primary_cards), primary_ui_count, limit_per_search):
        return SearchRunResult(
            search_term=search_term,
            time_filter=time_filter,
            cards=primary_cards,
            ui_observed_count=primary_ui_count,
            extracted_count=len(primary_cards),
            route_used="search-results",
            fallback_used=False,
        )

    fallback_cards, fallback_ui_count = _run_single_url(fallback_url)
    fallback_attempted = True
    combined: list[LinkedInJobCard] = []
    seen: set[str] = set()
    for card in [*primary_cards, *fallback_cards]:
        if card.url in seen:
            continue
        seen.add(card.url)
        combined.append(card)
        if limit_per_search is not None and len(combined) >= limit_per_search:
            break
    observed_counts = [count for count in (primary_ui_count, fallback_ui_count) if count is not None]
    return SearchRunResult(
        search_term=search_term,
        time_filter=time_filter,
        cards=combined,
        ui_observed_count=max(observed_counts) if observed_counts else None,
        extracted_count=len(combined),
        route_used="search-results+fallback" if fallback_attempted else "search-results",
        fallback_used=fallback_attempted,
    )


def cards_to_jobs(cards: Iterable[LinkedInJobCard]) -> list[dict]:
    jobs: list[dict] = []
    today = datetime.now().strftime("%Y-%m-%d")
    for card in cards:
        jobs.append(
            {
                "id": None,
                "date_found": today,
                "date_posted": _parse_relative_date(card.listed_at),
                "company": card.company,
                "role_title": card.title,
                "role_type": "Other",
                "location": card.location,
                "url": card.url,
                "url_hash": url_hash(card.url),
                "source": SOURCE_TAG,
                "fit_score": None,
                "fit_rationale": None,
                "status": "new",
                "date_applied": None,
                "folder_path": None,
                "jd_text": card.jd_text,
                "notes": (
                    f"tag={SOURCE_TAG} "
                    f"search={card.search_term} "
                    f"window={TIME_LABELS.get(card.time_filter, card.time_filter)} "
                    f"insight={card.insight}"
                ).strip(),
                "tc_hash": title_company_hash(card.title, card.company),
            }
        )
    return jobs


def append_new_jobs(df_existing: pd.DataFrame, jobs: list[dict]) -> tuple[pd.DataFrame, list[dict]]:
    existing_url_hashes = {
        str(v).strip()
        for v in df_existing.get("url_hash", pd.Series(dtype=str)).fillna("").tolist()
        if str(v).strip()
    }
    existing_tc_hashes = {
        title_company_hash(str(row.get("role_title") or ""), str(row.get("company") or ""))
        for _, row in df_existing.iterrows()
    }

    fresh: list[dict] = []
    seen_new_urls: set[str] = set()
    seen_new_tcs: set[str] = set()
    for job in jobs:
        if job["url_hash"] in existing_url_hashes or job["url_hash"] in seen_new_urls:
            continue
        if job["tc_hash"] in existing_tc_hashes or job["tc_hash"] in seen_new_tcs:
            continue
        fresh.append(job)
        seen_new_urls.add(job["url_hash"])
        seen_new_tcs.add(job["tc_hash"])

    if not fresh:
        return df_existing, []

    start_id = 1
    if not df_existing.empty and "id" in df_existing.columns:
        numeric_ids = pd.to_numeric(df_existing["id"], errors="coerce").dropna()
        if not numeric_ids.empty:
            start_id = int(numeric_ids.max()) + 1

    rows_for_df = []
    for idx, job in enumerate(fresh, start=start_id):
        job["id"] = str(idx)
        rows_for_df.append({col: job.get(col) for col in COLUMNS})

    df_new = pd.DataFrame(rows_for_df, columns=COLUMNS)
    if df_existing.empty:
        return df_new, fresh
    return pd.concat([df_existing[COLUMNS], df_new], ignore_index=True), fresh


def filter_jobs_for_write(jobs: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Keep only jobs we would actually want in the tracker.
    We do not persist:
    - rows with no JD text
    - scorer errors
    - hard rejects
    - deprioritized jobs
    """
    accepted: list[dict] = []
    dropped: list[dict] = []
    for job in jobs:
        has_jd = bool((job.get("jd_text") or "").strip())
        decision = str(job.get("decision") or "").strip().lower()
        if not has_jd or decision not in {"proceed"}:
            dropped.append(job)
            continue
        accepted.append(job)
    return accepted, dropped


def _run_post_extract_pipeline(
    *,
    run_label: str,
    searches: list[tuple[str, str]],
    search_runs: list[dict],
    jobs: list[dict],
    extracted_count: int,
    dry_run: bool,
    model: str,
    quiet: bool,
    max_workers: int,
    source_raw_artifacts: list[str] | None = None,
) -> int:
    df_existing = load_jobs()
    jobs_unseen, existing_hits = _split_existing_jobs(jobs, df_existing)
    review_cache = _load_review_cache()
    jobs_to_score, cache_hits = _split_cached_review_jobs(jobs_unseen, review_cache)

    if not quiet:
        print(f"Skipped as existing rows before scoring: {len(existing_hits)}")
        print(f"Skipped from review cache before scoring: {len(cache_hits)}")

    scored_new = score_batch(jobs_to_score, model=model, verbose=not quiet, max_workers=max_workers) if jobs_to_score else []
    reviewed_jobs = [*scored_new, *cache_hits]
    all_reviewed_jobs = [
        *reviewed_jobs,
        *[{**job, "__report_gate": "existing"} for job in existing_hits],
    ]
    accepted_for_write, dropped_before_write = filter_jobs_for_write(reviewed_jobs)
    merged_df, fresh = append_new_jobs(df_existing, accepted_for_write)
    cache_rows = _terminal_cache_rows(scored_new)

    score_artifact = _write_run_artifact(
        "linkedin_live_scored",
        {
            "extracted": extracted_count,
            "scored": len(scored_new),
            "reviewed": len(reviewed_jobs),
            "existing_skipped": len(existing_hits),
            "cache_skipped": len(cache_hits),
            "accepted_for_write": len(accepted_for_write),
            "dropped_before_write": len(dropped_before_write),
            "new_after_dedup": len(fresh),
            "searches": [{"search_term": s, "time_filter": t} for s, t in searches],
            "search_runs": search_runs,
            "source_raw_artifacts": source_raw_artifacts or [],
            "jobs": reviewed_jobs,
        },
    )
    if not quiet:
        print(f"Scored artifact: {score_artifact}")
        print(f"Scored fresh this run: {len(scored_new)}")
        print(f"Accepted for write after gating: {len(accepted_for_write)}")
        print(f"Dropped before write: {len(dropped_before_write)}")
        print(f"New jobs after dedup: {len(fresh)}")

    md_report, html_report = _write_batch_report(
        run_label=run_label,
        searches=searches,
        search_runs=search_runs,
        scored_jobs=reviewed_jobs,
        all_reviewed_jobs=all_reviewed_jobs,
        accepted_for_write=accepted_for_write,
        fresh_after_dedup=fresh,
        extracted_count=extracted_count,
        scored_count=len(scored_new),
        existing_skip_count=len(existing_hits),
        cache_hits_count=len(cache_hits),
    )
    if not quiet:
        print(f"Batch report (Markdown): {md_report}")
        print(f"Batch report (HTML): {html_report}")

    if fresh:
        save_jobs(merged_df, dry_run=dry_run)
        if dry_run:
            print(f"[dry-run] Would append {len(fresh)} LinkedIn-live jobs to {JOBS_XLSX}")
        else:
            print(f"Appended {len(fresh)} LinkedIn-live jobs to {JOBS_XLSX}")
    else:
        print("All captured jobs were duplicates of existing rows.")

    if not dry_run and cache_rows:
        _write_review_cache(cache_rows)
        if not quiet:
            print(f"Updated review cache with {len(cache_rows)} terminal decisions in {JOBS_XLSX} [{REVIEW_CACHE_SHEET_NAME}]")

    if not dry_run:
        run_bundle = _export_run_bundle(
            run_label=run_label,
            searches=searches,
            search_runs=search_runs,
            reviewed_jobs=reviewed_jobs,
            fresh_after_dedup=fresh,
            markdown_report=md_report,
            html_report=html_report,
            extracted_count=extracted_count,
            scored_count=len(scored_new),
            existing_skip_count=len(existing_hits),
            cache_hits_count=len(cache_hits),
        )
        if not quiet:
            print(f"Run bundle: {run_bundle}")

    return len(fresh)


def _searches_from_payload(payload: dict) -> list[tuple[str, str]]:
    searches: list[tuple[str, str]] = []
    for item in payload.get("searches", []):
        search_term = str(item.get("search_term") or "").strip()
        time_filter = str(item.get("time_filter") or "").strip()
        if search_term and time_filter:
            searches.append((search_term, time_filter))
    return searches


def _cards_from_payload(payload: dict) -> list[LinkedInJobCard]:
    cards: list[LinkedInJobCard] = []
    for item in payload.get("cards", []):
        try:
            cards.append(LinkedInJobCard(**item))
        except TypeError:
            continue
    return cards


def _dedupe_jobs_for_replay(jobs: Iterable[dict]) -> list[dict]:
    deduped: list[dict] = []
    seen_urls: set[str] = set()
    seen_tcs: set[str] = set()
    for job in jobs:
        url_key = str(job.get("url_hash") or "").strip()
        tc_key = str(job.get("tc_hash") or "").strip()
        if url_key and url_key in seen_urls:
            continue
        if tc_key and tc_key in seen_tcs:
            continue
        if url_key:
            seen_urls.add(url_key)
        if tc_key:
            seen_tcs.add(tc_key)
        deduped.append(job)
    return deduped


def score_from_raw_artifacts(
    raw_artifact_paths: list[str],
    dry_run: bool,
    model: str,
    quiet: bool,
    max_workers: int = 2,
) -> int:
    payloads: list[tuple[Path, dict]] = []
    for raw_path in raw_artifact_paths:
        path = Path(raw_path).expanduser().resolve()
        payload = json.loads(path.read_text(encoding="utf-8"))
        payloads.append((path, payload))

    searches: list[tuple[str, str]] = []
    search_runs: list[dict] = []
    replay_jobs: list[dict] = []
    extracted_count = 0

    for path, payload in payloads:
        searches.extend(_searches_from_payload(payload))
        search_runs.extend(payload.get("search_runs", []))
        jobs = payload.get("jobs")
        if isinstance(jobs, list) and jobs:
            replay_jobs.extend(dict(job) for job in jobs)
            extracted_count += int(payload.get("count") or len(jobs))
        else:
            cards = _cards_from_payload(payload)
            replay_jobs.extend(cards_to_jobs(cards))
            extracted_count += int(payload.get("count") or len(cards))

    replay_jobs = _dedupe_jobs_for_replay(replay_jobs)
    if not replay_jobs:
        print("No jobs found in the provided raw artifact(s).")
        return 0

    windows = sorted({TIME_LABELS.get(window, window) for _, window in searches}) if searches else ["raw_replay"]
    run_label = windows[0] if len(windows) == 1 else "mixed"

    if not quiet:
        print(f"Loaded {len(replay_jobs)} unique jobs from {len(payloads)} raw artifact(s).")
        for path, _ in payloads:
            print(f"  Raw input: {path}")

    return _run_post_extract_pipeline(
        run_label=run_label,
        searches=searches,
        search_runs=search_runs,
        jobs=replay_jobs,
        extracted_count=extracted_count,
        dry_run=dry_run,
        model=model,
        quiet=quiet,
        max_workers=max_workers,
        source_raw_artifacts=[str(path) for path, _ in payloads],
    )


def run_live_discovery(
    searches: list[tuple[str, str]],
    debug_port: int,
    limit_per_search: int | None,
    pages: int | None,
    dry_run: bool,
    model: str,
    quiet: bool,
    extract_only: bool = False,
    count_only: bool = False,
    max_workers: int = 2,
) -> int:
    windows = sorted({TIME_LABELS.get(window, window) for _, window in searches})
    run_label = windows[0] if len(windows) == 1 else "mixed"
    run_stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    progress_path, raw_inflight_path = _inflight_artifact_paths(run_stamp, run_label)
    search_run_summaries: list[dict] = []
    scraped_cards: list[LinkedInJobCard] = []
    current_search_cards: list[LinkedInJobCard] = []
    repair_summary = JdRepairSummary(candidates=0, attempted=0, repaired=0, remaining_failed=0)
    progress_state: dict[str, object] = {
        "run_label": run_label,
        "run_stamp": run_stamp,
        "status": "starting",
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "last_progress_at": None,
        "last_heartbeat_at": None,
        "event": "starting",
        "searches": [{"search_term": term, "time_filter": window} for term, window in searches],
        "current_search_term": "",
        "current_time_filter": "",
        "current_route": "",
        "current_page_index": 0,
        "current_start": 0,
        "current_search_extracted": 0,
        "total_extracted": 0,
        "ui_observed_count": None,
        "searches_completed": 0,
        "inflight_raw_path": str(raw_inflight_path),
    }
    last_progress_write = 0.0
    last_partial_raw_write = 0.0

    def _partial_raw_payload() -> dict:
        combined_cards = [*scraped_cards, *current_search_cards]
        return {
            "status": progress_state.get("status"),
            "run_label": run_label,
            "run_stamp": run_stamp,
            "count": len(combined_cards),
            "searches": [{"search_term": s, "time_filter": t} for s, t in searches],
            "search_runs": search_run_summaries,
            "progress": progress_state,
            "cards": [asdict(card) for card in combined_cards],
            "jobs": cards_to_jobs(combined_cards),
        }

    def _emit_progress(
        *,
        force: bool = False,
        event: str,
        status: str | None = None,
        search_term: str | None = None,
        time_filter: str | None = None,
        route: str | None = None,
        page_index: int | None = None,
        start: int | None = None,
        search_extracted: int | None = None,
        total_extracted: int | None = None,
        ui_observed_count: int | None = None,
        cards_snapshot: list[LinkedInJobCard] | None = None,
        error: str | None = None,
        progress_made: bool = False,
    ) -> None:
        nonlocal last_progress_write, last_partial_raw_write, current_search_cards
        now_mono = time.monotonic()
        now_iso = datetime.now().isoformat(timespec="seconds")

        if status is not None:
            progress_state["status"] = status
        progress_state["event"] = event
        if search_term is not None:
            progress_state["current_search_term"] = search_term
        if time_filter is not None:
            progress_state["current_time_filter"] = time_filter
        if route is not None:
            progress_state["current_route"] = route
        if page_index is not None:
            progress_state["current_page_index"] = page_index
        if start is not None:
            progress_state["current_start"] = start
        if search_extracted is not None:
            progress_state["current_search_extracted"] = search_extracted
        if total_extracted is not None:
            progress_state["total_extracted"] = total_extracted
        if ui_observed_count is not None:
            progress_state["ui_observed_count"] = ui_observed_count
        if error is not None:
            progress_state["error"] = error
        if cards_snapshot is not None:
            current_search_cards = list(cards_snapshot)
        if progress_made:
            progress_state["last_progress_at"] = now_iso

        should_write = force or (now_mono - last_progress_write) >= HEARTBEAT_INTERVAL_SECONDS
        if not should_write:
            return

        progress_state["last_heartbeat_at"] = now_iso
        _write_json_artifact(progress_path, progress_state)
        if force or (now_mono - last_partial_raw_write) >= HEARTBEAT_INTERVAL_SECONDS:
            _write_json_artifact(raw_inflight_path, _partial_raw_payload())
            last_partial_raw_write = now_mono
        last_progress_write = now_mono

        if not quiet:
            window_label = TIME_LABELS.get(str(progress_state.get("current_time_filter") or ""), str(progress_state.get("current_time_filter") or ""))
            print(
                "[heartbeat] "
                f"status={progress_state.get('status')} "
                f"search={progress_state.get('current_search_term') or '-'} "
                f"window={window_label or '-'} "
                f"route={progress_state.get('current_route') or '-'} "
                f"page={progress_state.get('current_page_index')} "
                f"search_count={progress_state.get('current_search_extracted')} "
                f"total_count={progress_state.get('total_extracted')}"
            )

    _emit_progress(force=True, event="starting", status="starting", total_extracted=0)

    try:
        with sync_playwright() as playwright:
            session = _open_linkedin_browser_session(playwright, debug_port)
            page: Page | None = None
            detail_page: Page | None = None
            try:
                context = session["context"]
                preflight = _session_preflight(context)
                if not preflight.get("ok"):
                    raise RuntimeError(
                        "LinkedIn live preflight failed. "
                        f"URL={preflight.get('current_url', '')} "
                        f"title={preflight.get('title', '')} "
                        f"authwall_or_login={preflight.get('authwall_or_login')} "
                        f"has_li_at_cookie={preflight.get('has_li_at_cookie')}"
                    )
                if not quiet:
                    print(
                        "LinkedIn preflight OK: "
                        f"url={preflight.get('current_url', '')} "
                        f"pages_before={preflight.get('context_pages_before')}"
                    )

                _emit_progress(force=True, event="preflight_ok", status="running", total_extracted=0, progress_made=True)

                page = context.new_page()
                detail_page = context.new_page()
                page.set_default_timeout(15000)
                detail_page.set_default_timeout(15000)

                for search_idx, (search_term, time_filter) in enumerate(searches, start=1):
                    current_search_cards = []
                    _emit_progress(
                        force=True,
                        event="search_started",
                        status="running",
                        search_term=search_term,
                        time_filter=time_filter,
                        page_index=0,
                        start=0,
                        search_extracted=0,
                        total_extracted=len(scraped_cards),
                        progress_made=True,
                    )
                    if not quiet:
                        print(f"\nSearching LinkedIn Jobs: {search_term} | {TIME_LABELS.get(time_filter, time_filter)}")
                    result = scrape_search(
                        page=page,
                        search_term=search_term,
                        time_filter=time_filter,
                        limit_per_search=limit_per_search,
                        pages=pages,
                        count_only=count_only,
                        detail_page=detail_page,
                        progress_callback=lambda payload: _emit_progress(
                            event=str(payload.get("event") or "progress"),
                            status="running",
                            search_term=str(payload.get("search_term") or search_term),
                            time_filter=str(payload.get("time_filter") or time_filter),
                            route=str(payload.get("route") or ""),
                            page_index=int(payload.get("page_index") or 0),
                            start=int(payload.get("start") or 0),
                            search_extracted=int(payload.get("search_extracted") or 0),
                            total_extracted=len(scraped_cards) + int(payload.get("search_extracted") or 0),
                            ui_observed_count=payload.get("ui_observed_count"),
                            cards_snapshot=payload.get("cards_snapshot"),
                            progress_made=str(payload.get("event") or "") == "new_cards",
                        ),
                    )
                    cards = result.cards
                    current_search_cards = cards
                    search_run_summaries.append(
                        {
                            "search_term": result.search_term,
                            "time_filter": result.time_filter,
                            "ui_observed_count": result.ui_observed_count,
                            "extracted_count": result.extracted_count,
                            "route_used": result.route_used,
                            "fallback_used": result.fallback_used,
                            "count_mismatch": (
                                result.ui_observed_count is not None
                                and result.ui_observed_count != result.extracted_count
                            ),
                        }
                    )
                    scraped_cards.extend(cards)
                    progress_state["searches_completed"] = search_idx
                    current_search_cards = []
                    _emit_progress(
                        force=True,
                        event="search_complete",
                        status="running",
                        search_term=search_term,
                        time_filter=time_filter,
                        route=result.route_used,
                        search_extracted=len(cards),
                        total_extracted=len(scraped_cards),
                        ui_observed_count=result.ui_observed_count,
                        progress_made=True,
                    )
                    if not quiet:
                        ui_note = (
                            f" | UI reported total {result.ui_observed_count}"
                            if result.ui_observed_count is not None
                            else ""
                        )
                        print(f"  Captured {len(cards)} cards{ui_note}")

                if not count_only and scraped_cards:
                    _emit_progress(
                        force=True,
                        event="jd_repair_started",
                        status="repairing",
                        total_extracted=len(scraped_cards),
                        progress_made=True,
                    )
                    repair_summary = _repair_scraped_cards(
                        page=page,
                        detail_page=detail_page,
                        cards=scraped_cards,
                        quiet=quiet,
                        progress_callback=lambda payload: _emit_progress(
                            force=str(payload.get("event") or "").endswith("success"),
                            event=str(payload.get("event") or "jd_repair_progress"),
                            status="repairing",
                            total_extracted=len(scraped_cards),
                            progress_made=str(payload.get("event") or "") == "jd_repair_success",
                        ),
                    )
                    _emit_progress(
                        force=True,
                        event="jd_repair_complete",
                        status="running",
                        total_extracted=len(scraped_cards),
                        progress_made=repair_summary.repaired > 0,
                    )
            finally:
                _close_page_safely(detail_page)
                _close_page_safely(page)
                try:
                    session["cleanup"]()
                except Exception:
                    pass
    except Exception as exc:
        _emit_progress(
            force=True,
            event="failed",
            status="failed",
            total_extracted=len(scraped_cards) + len(current_search_cards),
            error=str(exc),
        )
        raise

    jobs = cards_to_jobs(scraped_cards)
    _emit_progress(force=True, event="extraction_complete", status="extraction_complete", total_extracted=len(scraped_cards), progress_made=True)
    artifact = _write_run_artifact(
        "linkedin_live_raw",
        {
            "count": len(scraped_cards),
            "searches": [{"search_term": s, "time_filter": t} for s, t in searches],
            "search_runs": search_run_summaries,
            "jd_repair_summary": asdict(repair_summary),
            "cards": [asdict(card) for card in scraped_cards],
            "jobs": jobs,
        },
    )

    if not quiet:
        print(f"\nRaw artifact: {artifact}")
        print(f"Raw cards captured: {len(scraped_cards)}")
        if repair_summary.attempted:
            print(
                "JD repair summary: "
                f"candidates={repair_summary.candidates} "
                f"attempted={repair_summary.attempted} "
                f"repaired={repair_summary.repaired} "
                f"remaining_failed={repair_summary.remaining_failed}"
            )

    if not jobs:
        _emit_progress(force=True, event="complete", status="complete", total_extracted=0)
        print("No jobs were captured from LinkedIn live search.")
        return 0

    if extract_only:
        md_report, html_report = _write_batch_report(
            run_label=f"{run_label}_{'count_only' if count_only else 'extract_only'}",
            searches=searches,
            search_runs=search_run_summaries,
            scored_jobs=jobs,
            all_reviewed_jobs=jobs,
            accepted_for_write=[],
            fresh_after_dedup=[],
            extracted_count=len(jobs),
            scored_count=0,
            existing_skip_count=0,
            cache_hits_count=0,
        )
        _emit_progress(force=True, event="complete", status="complete", total_extracted=len(jobs), progress_made=True)
        print(f"{'Count-only' if count_only else 'Extract-only'} complete: {len(jobs)} jobs")
        print(f"Batch report (Markdown): {md_report}")
        print(f"Batch report (HTML): {html_report}")
        return len(jobs)
    _emit_progress(force=True, event="scoring_started", status="scoring", total_extracted=len(jobs), progress_made=True)
    try:
        fresh_count = _run_post_extract_pipeline(
            run_label=run_label,
            searches=searches,
            search_runs=search_run_summaries,
            jobs=jobs,
            extracted_count=len(scraped_cards),
            dry_run=dry_run,
            model=model,
            quiet=quiet,
            max_workers=max_workers,
            source_raw_artifacts=[str(artifact)],
        )
    except Exception as exc:
        _emit_progress(force=True, event="failed", status="failed", total_extracted=len(jobs), error=str(exc))
        raise
    _emit_progress(force=True, event="complete", status="complete", total_extracted=len(jobs), progress_made=True)
    return fresh_count


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Logged-in LinkedIn live discovery")
    parser.add_argument("--debug-port", type=int, default=DEFAULT_DEBUG_PORT, help="Chrome remote debugging port")
    parser.add_argument(
        "--limit-per-search",
        type=int,
        default=DEFAULT_LIMIT_PER_SEARCH,
        help="Max jobs to inspect per search. Omit for no cap. Use 0 for no cap.",
    )
    parser.add_argument(
        "--pages",
        type=int,
        default=DEFAULT_PAGES,
        help="LinkedIn result pages to scan per search. Omit for all pages. Use 0 for all pages.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Score jobs but skip writing jobs.xlsx")
    parser.add_argument("--model", type=str, default=DEFAULT_MODEL, help=f"Scoring model (default: {DEFAULT_MODEL})")
    parser.add_argument("--extract-only", action="store_true", help="Only extract/count jobs. Skip JD scoring and jobs.xlsx writes.")
    parser.add_argument(
        "--count-only",
        action="store_true",
        help="Capture counts/cards only. Do not open job details for JD extraction.",
    )
    parser.add_argument(
        "--score-from-raw",
        action="append",
        default=[],
        help="Replay scoring/write from a saved linkedin_live_raw_*.json artifact. Repeat for multiple artifacts.",
    )
    parser.add_argument("--max-workers", type=int, default=2, help="Parallel scoring workers for the full scored run.")
    parser.add_argument("--quiet", action="store_true", help="Reduce terminal output")
    parser.add_argument("--search", action="append", default=[], help="Override search term. Repeat for multiple values.")
    parser.add_argument(
        "--time",
        action="append",
        default=[],
        help="Override LinkedIn recency filter(s), e.g. r86400 or r604800. Must align with --search count if provided.",
    )
    return parser.parse_args()


def _resolve_searches(args: argparse.Namespace) -> list[tuple[str, str]]:
    if not args.search and not args.time:
        return list(DEFAULT_SEARCHES)
    if len(args.search) != len(args.time):
        raise SystemExit("When overriding searches, provide matching counts for --search and --time.")
    return list(zip(args.search, args.time))


if __name__ == "__main__":
    args = _parse_args()
    if args.score_from_raw:
        score_from_raw_artifacts(
            raw_artifact_paths=args.score_from_raw,
            dry_run=args.dry_run,
            model=args.model,
            quiet=args.quiet,
            max_workers=args.max_workers,
        )
    else:
        if args.count_only and not args.extract_only:
            raise SystemExit("--count-only requires --extract-only.")
        searches = _resolve_searches(args)
        limit_per_search = None if args.limit_per_search in (None, 0) else args.limit_per_search
        pages = None if args.pages in (None, 0) else args.pages
        run_live_discovery(
            searches=searches,
            debug_port=args.debug_port,
            limit_per_search=limit_per_search,
            pages=pages,
            dry_run=args.dry_run,
            model=args.model,
            quiet=args.quiet,
            extract_only=args.extract_only,
            count_only=args.count_only,
            max_workers=args.max_workers,
        )
