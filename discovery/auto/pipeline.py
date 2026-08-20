"""
pipeline.py — Stage 1 Orchestrator
-------------------------------------
Ties scraper → scorer → jobs.xlsx into a single run.

What it does each run:
  1. Load existing url_hashes from jobs.xlsx (for dedup)
  2. Scrape new jobs via JobSpy (Lane A + Lane B query packs)
  3. Score each new job via Claude
  4. Write results to jobs.xlsx (append, never overwrite existing rows)
  5. Print a run digest

Usage:
    python discovery/auto/pipeline.py                  # full run
    python discovery/auto/pipeline.py --dry-run        # scrape + score, don't write xlsx
    python discovery/auto/pipeline.py --skip-scrape    # score jobs already in xlsx with status=new
    python discovery/auto/pipeline.py --hours-old 48   # widen scraper lookback window

Run from ResumeGenerator v1/ root.
"""

import argparse
import hashlib
import io
import os
import re
import subprocess
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_HERE      = Path(__file__).parent
_ROOT      = _HERE.parent
JOBS_XLSX  = _ROOT / "jobs.xlsx"
LOGS_DIR   = _HERE / "logs"
SHEET_NAME = "Jobs"


def _wait_process_wall_clock(process: subprocess.Popen, timeout_seconds: float) -> int:
    """Wait using epoch time so machine sleep counts toward the hard run cap."""
    deadline_epoch = time.time() + max(timeout_seconds, 0.1)
    while True:
        remaining = deadline_epoch - time.time()
        if remaining <= 0:
            raise subprocess.TimeoutExpired(process.args, timeout_seconds)
        try:
            return process.wait(timeout=min(1.0, remaining))
        except subprocess.TimeoutExpired:
            continue

# ---------------------------------------------------------------------------
# xlsx helpers
# ---------------------------------------------------------------------------

# Canonical column order — must match jobs.xlsx schema
COLUMNS = [
    "id", "date_found", "date_posted", "company", "role_title", "role_type",
    "location", "url", "url_hash", "source",
    "fit_score", "fit_rationale", "status",
    "date_applied", "folder_path", "resume_run", "jd_text", "notes",
    "lane", "deadline", "deadline_source", "everify_status",
    "sponsorship_flag", "classification", "reject_reason",
]

# Columns written by scraper+scorer (tc_hash is internal, not stored)
_INTERNAL_COLS = {"tc_hash", "_query_id", "_raw_response", "breakdown",
                  "decision", "category", "lane", "start_timing",
                  "application_deadline", "deadline_source", "deadline_lookup",
                  "e_verify_status", "eligibility_flags", "hourly_rate_min",
                  "hourly_rate_max", "rejection_reason", "discovery_disposition",
                  "discovery_reason", "role_family", "query_lane",
                  "_scoring_rate_limit_events", "_run_timeout_unscored"}

# ---------------------------------------------------------------------------
# Formatting constants (keep in sync with jobs.py)
# ---------------------------------------------------------------------------
BLOCKLIST_PATH = _ROOT / "blocklist.txt"   # discovery/blocklist.txt

_STATUS_RANK = {
    "queued":    0,
    "review":    1,
    "promoted":  2,
    "generated": 3,
    "applied":   4,
    "new":       5,
    "skipped":   6,
    "skip":      6,
    "rejected":  6,
}

_STATUS_COLOR = {
    "queued":    "E2EFDA",
    "review":    "FFF2CC",
    "promoted":  "BDD7EE",
    "generated": "DDEBF7",
    "applied":   "EDEDED",
    "new":       None,
    "skipped":   "C0C0C0",
    "skip":      "C0C0C0",
    "rejected":  "C0C0C0",
}

_SCORE_COLOR = {
    "high": ("C6EFCE", "276221"),
    "mid":  ("FFEB9C", "9C6500"),
    "low":  ("FFC7CE", "9C0006"),
}

_COL_WIDTHS = {
    "id": 6, "date_found": 13, "date_posted": 13, "company": 22,
    "role_title": 32, "role_type": 12, "location": 16, "url": 40,
    "url_hash": 14, "source": 12, "fit_score": 10, "fit_rationale": 50,
    "status": 12, "date_applied": 13, "folder_path": 40, "resume_run": 12, "jd_text": 18,
    "notes": 40, "lane": 8, "deadline": 22, "deadline_source": 18,
    "everify_status": 16, "sponsorship_flag": 28, "classification": 14,
    "reject_reason": 50,
}


def _safe_float_p(val) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _load_blocklist_p() -> list[str]:
    if not BLOCKLIST_PATH.exists():
        return []
    lines = []
    for ln in BLOCKLIST_PATH.read_text(encoding="utf-8").splitlines():
        ln = ln.split("#")[0].strip()
        if ln:
            lines.append(ln.lower())
    return lines


def _is_blocklisted_p(company: str, patterns: list[str]) -> bool:
    import fnmatch
    co = company.lower().strip()
    return any(fnmatch.fnmatch(co, p) for p in patterns)


def _sort_df_p(df: pd.DataFrame) -> pd.DataFrame:
    """Sort df by STATUS_RANK, then fit_score desc within each group."""
    bl = _load_blocklist_p()
    df = df.copy()
    df["_rank"]    = df["status"].apply(
        lambda s: _STATUS_RANK.get(str(s).lower().strip(), 99)
    )
    df["_score"]   = df["fit_score"].apply(_safe_float_p)
    df["_blocked"] = df["company"].apply(
        lambda co: _is_blocklisted_p(str(co), bl)
    )
    df["_date_applied"] = pd.to_datetime(df.get("date_applied", ""), errors="coerce")
    groups = []
    for rank in sorted(df["_rank"].unique()):
        grp = df[df["_rank"] == rank].copy()
        if rank == _STATUS_RANK["applied"]:
            grp = grp.sort_values(
                ["_blocked", "_date_applied"], ascending=[True, False], na_position="last"
            )
        else:
            grp = grp.sort_values(["_blocked", "_score"], ascending=[True, False])
        groups.append(grp)
    df = pd.concat(groups) if groups else df
    return df.drop(columns=["_rank", "_score", "_blocked", "_date_applied"])


def load_jobs() -> pd.DataFrame:
    """Load jobs.xlsx. Returns empty DataFrame with correct columns if file missing."""
    if not JOBS_XLSX.exists():
        return pd.DataFrame(columns=COLUMNS)
    try:
        df = pd.read_excel(JOBS_XLSX, sheet_name=SHEET_NAME, dtype=str)
        # Ensure all expected columns exist
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = None
        return df[COLUMNS]
    except Exception as e:
        print(f"  ⚠  Could not read jobs.xlsx: {e}")
        return pd.DataFrame(columns=COLUMNS)


def save_jobs(df: pd.DataFrame, dry_run: bool = False) -> None:
    """
    Sort, write, and fully format jobs.xlsx.
    Applies status colours, score traffic lights, section dividers, and
    blocklist strikethrough — same visual spec as jobs.py save_jobs().
    (Keep in sync with jobs.py save_jobs() if you change the style.)
    """
    if dry_run:
        print("  [dry-run] Would write to jobs.xlsx — skipped")
        return

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Sort before writing so section dividers land in the right places
    df = _sort_df_p(df)

    if JOBS_XLSX.exists():
        with pd.ExcelWriter(
            JOBS_XLSX, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
    else:
        with pd.ExcelWriter(JOBS_XLSX, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

    wb = load_workbook(JOBS_XLSX)
    ws = wb[SHEET_NAME]

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_mid    = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    col_map = {col: i + 1 for i, col in enumerate(df.columns)}

    for col, col_i in col_map.items():
        cell = ws.cell(row=1, column=col_i)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_wrap
        ws.column_dimensions[get_column_letter(col_i)].width = _COL_WIDTHS.get(col, 18)

    ws.freeze_panes      = "A2"
    ws.row_dimensions[1].height = 28
    ws.auto_filter.ref   = ws.dimensions

    # ── Data rows ─────────────────────────────────────────────────────────────
    divider_border = Border(top=Side(border_style="medium", color="1F3864"))
    base_font      = Font(size=10, name="Calibri")
    blocked_fill   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    blocked_font   = Font(color="AAAAAA", italic=True, size=10, name="Calibri", strike=True)

    status_col_i  = col_map.get("status")
    score_col_i   = col_map.get("fit_score")
    bl_patterns   = _load_blocklist_p()
    prev_rank     = None

    for row_i, (_, row) in enumerate(df.iterrows(), start=2):
        status     = str(row.get("status", "")).lower().strip()
        score      = _safe_float_p(row.get("fit_score"))
        rank       = _STATUS_RANK.get(status, 99)
        is_blocked = _is_blocklisted_p(str(row.get("company", "")), bl_patterns)

        # Thick divider between status groups
        if prev_rank is not None and rank != prev_rank:
            for col_i in range(1, len(df.columns) + 1):
                ws.cell(row=row_i, column=col_i).border = divider_border

        ws.row_dimensions[row_i].height = 15
        for col_i in range(1, len(df.columns) + 1):
            ws.cell(row=row_i, column=col_i).alignment = left_mid
            ws.cell(row=row_i, column=col_i).font      = base_font

        if is_blocked:
            for col_i in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_i, column=col_i)
                cell.fill = blocked_fill
                cell.font = blocked_font
        else:
            if status_col_i:
                hex_color = _STATUS_COLOR.get(status)
                sc = ws.cell(row=row_i, column=status_col_i)
                if hex_color:
                    sc.fill = PatternFill(start_color=hex_color, end_color=hex_color,
                                          fill_type="solid")
                sc.font = Font(bold=True, size=10, name="Calibri")

            if score_col_i and score > 0:
                tier = "high" if score >= 8.5 else ("mid" if score >= 7.0 else "low")
                fill_hex, font_hex = _SCORE_COLOR[tier]
                sc = ws.cell(row=row_i, column=score_col_i)
                sc.fill = PatternFill(start_color=fill_hex, end_color=fill_hex,
                                      fill_type="solid")
                sc.font = Font(color=font_hex, bold=True, size=10, name="Calibri")

        prev_rank = rank

    wb.save(JOBS_XLSX)


def get_existing_hashes(df: pd.DataFrame) -> set[str]:
    """Return set of url_hashes already in the DataFrame."""
    return set(df["url_hash"].dropna().astype(str).tolist())


def _url_hash(url: str) -> str:
    return hashlib.md5(url.strip().lower().encode()).hexdigest()


def jobs_to_rows(jobs: list[dict], start_id: int) -> list[dict]:
    """
    Convert scored job dicts to clean row dicts matching COLUMNS schema.
    Strips internal-only fields, fills url_hash if missing, assigns IDs.
    """
    rows = []
    for i, job in enumerate(jobs):
        url = str(job.get("url") or "").strip()
        row = {col: job.get(col) for col in COLUMNS}
        decision = str(job.get("decision") or "").strip().lower()
        classification = str(
            job.get("classification") or job.get("discovery_disposition") or ""
        ).strip().lower()
        if decision in {"reject", "deprioritize"}:
            classification = "reject"
        elif decision == "unsure":
            classification = "unsure"
        elif decision == "proceed":
            classification = "keep"

        flags = job.get("eligibility_flags") or []
        if isinstance(flags, str):
            flags = [part.strip() for part in re.split(r"[,|]", flags) if part.strip()]
        sponsorship_flag = str(job.get("sponsorship_flag") or "").strip()
        if not sponsorship_flag and "h1b_sponsorship_unavailable" in flags:
            sponsorship_flag = "h1b_sponsorship_unavailable"

        row["lane"] = str(job.get("lane") or "").strip()
        row["deadline"] = job.get("deadline") or job.get("application_deadline") or ""
        row["deadline_source"] = job.get("deadline_source") or ""
        row["everify_status"] = job.get("everify_status") or job.get("e_verify_status") or ""
        row["sponsorship_flag"] = sponsorship_flag
        row["classification"] = classification
        row["reject_reason"] = (
            job.get("reject_reason")
            or job.get("rejection_reason")
            or (job.get("discovery_reason") if classification in {"reject", "unsure"} else "")
            or ""
        )
        row["id"]       = start_id + i
        row["url_hash"] = job.get("url_hash") or (_url_hash(url) if url else "")
        # Clean up None-ish values
        for k, v in row.items():
            if v is None or (isinstance(v, float) and pd.isna(v)):
                row[k] = ""
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Digest printer
# ---------------------------------------------------------------------------

def _watched_programs(new_jobs: list[dict]) -> dict[str, list[dict]]:
    google_apm: list[dict] = []
    meta_rpm: list[dict] = []
    for job in new_jobs:
        company = str(job.get("company") or "")
        title = str(job.get("role_title") or "")
        if re.search(r"\b(?:google|alphabet)\b", company, re.I) and re.search(
            r"\b(?:APM|associate\s+product\s+manager)\b", title, re.I
        ):
            google_apm.append(job)
        if re.search(r"\b(?:meta|facebook)\b", company, re.I) and re.search(
            r"\b(?:RPM|rotational\s+product\s+manager)\b", title, re.I
        ):
            meta_rpm.append(job)
    return {"Google APM": google_apm, "Meta RPM": meta_rpm}


def print_digest(new_jobs: list[dict], run_start: datetime,
                 scrape_report: dict | None = None) -> None:
    elapsed = (datetime.now() - run_start).seconds
    proceed = [j for j in new_jobs if j.get("decision") == "Proceed"]
    skipped = [j for j in new_jobs if j.get("decision") in ("Reject", "Deprioritize")]
    unsure  = [j for j in new_jobs if j.get("decision") == "Unsure"]
    errors  = [j for j in new_jobs if j.get("decision") == "Error"]
    high    = [j for j in proceed if j.get("category") == "High Priority"]
    mid     = [j for j in proceed if j.get("category") == "Medium Priority"]
    low     = [j for j in proceed if j.get("category") == "Low Priority"]

    print(f"\n{'═'*60}")
    print(f"  Run complete  ({elapsed}s)  —  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'═'*60}")
    print(f"  New jobs found:   {len(new_jobs)}")
    print(f"  Queued (proceed): {len(proceed)}  "
          f"[High: {len(high)}  Mid: {len(mid)}  Low: {len(low)}]")
    print(f"  Rejected/skipped: {len(skipped)}")
    print(f"  Unsure / review:  {len(unsure)}")
    print(f"  Scoring errors:   {len(errors)}")
    scoring_throttles = sum(int(j.get("_scoring_rate_limit_events") or 0) for j in new_jobs)
    print(f"  Scoring throttles:{scoring_throttles:>4}")

    if scrape_report:
        status_counts = scrape_report.get("query_status_counts") or {}
        print(
            f"  Query runtime:    {scrape_report.get('elapsed_seconds', 0):.1f}s | "
            f"completed={status_counts.get('completed', 0)} | "
            f"timed_out={status_counts.get('timed_out', 0)} | "
            f"failed={status_counts.get('failed', 0)} | "
            f"run-cap-skipped={status_counts.get('skipped_run_timeout', 0)}"
        )
        print(f"  Source throttles: {int(scrape_report.get('throttle_events') or 0)}")
        print(f"  Checkpoints:      {scrape_report.get('checkpoint_dir', '')}")

    print("\n  Volume by discovery lane:")
    for lane in ("A", "B", "C"):
        lane_jobs = [j for j in new_jobs if str(j.get("lane") or "A").upper() == lane]
        lane_proceed = sum(j.get("decision") == "Proceed" for j in lane_jobs)
        lane_reject = sum(j.get("decision") in ("Reject", "Deprioritize") for j in lane_jobs)
        lane_unsure = sum(j.get("decision") == "Unsure" for j in lane_jobs)
        lane_errors = sum(j.get("decision") == "Error" for j in lane_jobs)
        print(
            f"    Lane {lane}: {len(lane_jobs)} found | {lane_proceed} proceed | "
            f"{lane_reject} rejected/deprioritized | {lane_unsure} unsure | {lane_errors} errors"
        )

    rejection_reasons = Counter(_rejection_reason(j) for j in skipped)
    rejection_reasons.pop("", None)
    if rejection_reasons:
        print("\n  Rejects by reason:")
        for reason, count in rejection_reasons.most_common():
            print(f"    {count:>3}  {reason}")

    print("\n  Programme watch:")
    for label, matches in _watched_programs(new_jobs).items():
        print(f"    {label}: {len(matches)}")
        for job in matches:
            print(f"      {job.get('company')} — {job.get('role_title')} | {job.get('url', '')}")

    if high:
        print(f"\n  ★  Top picks this run:")
        top = sorted(high, key=lambda j: j.get("fit_score") or 0, reverse=True)
        for j in top[:5]:
            score = j.get("fit_score") or "?"
            print(f"     [{score}/10]  {j.get('company', '?')} — {j.get('role_title', '?')}")
            print(f"             {j.get('location', '')} | {j.get('source', '')}")
    print(f"{'═'*60}\n")


def _rejection_reason(job: dict) -> str:
    explicit = str(job.get("reject_reason") or job.get("rejection_reason") or "").strip()
    if explicit:
        return explicit
    rationale = str(job.get("fit_rationale") or "").strip()
    rationale = re.sub(r"^\[(?:Reject|Deprioritize)\s*\|[^]]*\]\s*", "", rationale, flags=re.I)
    return rationale or "Unspecified rejection/deprioritization"


# ---------------------------------------------------------------------------
# Run log writer
# ---------------------------------------------------------------------------

def write_run_log(new_jobs: list[dict], run_start: datetime,
                  hours_old: int, dry_run: bool,
                  scrape_report: dict | None = None) -> Path:
    """
    Write a structured log file for this run to discovery/auto/logs/.
    Returns the log file path.
    """
    LOGS_DIR.mkdir(exist_ok=True)
    timestamp   = run_start.strftime("%Y-%m-%d_%H%M")
    log_path    = LOGS_DIR / f"run_{timestamp}.txt"
    elapsed     = (datetime.now() - run_start).seconds

    proceed  = [j for j in new_jobs if j.get("decision") == "Proceed"]
    rejected = [j for j in new_jobs if j.get("decision") in ("Reject", "Deprioritize")]
    unsure   = [j for j in new_jobs if j.get("decision") == "Unsure"]
    errors   = [j for j in new_jobs if j.get("decision") == "Error"]
    high     = sorted([j for j in proceed if j.get("category") == "High Priority"],
                      key=lambda j: j.get("fit_score") or 0, reverse=True)
    mid      = [j for j in proceed if j.get("category") == "Medium Priority"]
    low      = [j for j in proceed if j.get("category") == "Low Priority"]

    lines = [
        f"Pipeline Run Log",
        f"{'='*60}",
        f"Run time:    {run_start.strftime('%Y-%m-%d %H:%M')}",
        f"Elapsed:     {elapsed}s",
        f"Lookback:    {hours_old}h",
        f"Dry run:     {dry_run}",
        f"",
        f"── Summary ──────────────────────────────────────────────",
        f"New jobs found:    {len(new_jobs)}",
        f"Queued (proceed):  {len(proceed)}  "
            f"[High: {len(high)}  Mid: {len(mid)}  Low: {len(low)}]",
        f"Rejected/skipped:  {len(rejected)}",
        f"Unsure / review:   {len(unsure)}",
        f"Scoring errors:    {len(errors)}",
        f"Scoring throttles: {sum(int(j.get('_scoring_rate_limit_events') or 0) for j in new_jobs)}",
        f"",
    ]

    if scrape_report:
        status_counts = scrape_report.get("query_status_counts") or {}
        lines += [
            f"── Scrape execution ──────────────────────────────────────",
            f"Runtime:             {scrape_report.get('elapsed_seconds', 0):.1f}s",
            f"Completed queries:    {status_counts.get('completed', 0)}",
            f"Timed-out queries:    {status_counts.get('timed_out', 0)}",
            f"Failed queries:       {status_counts.get('failed', 0)}",
            f"Run-cap skipped:      {status_counts.get('skipped_run_timeout', 0)}",
            f"Source throttles:     {int(scrape_report.get('throttle_events') or 0)}",
            f"Checkpoint directory: {scrape_report.get('checkpoint_dir', '')}",
            "",
        ]
        problem_queries = [
            record for record in scrape_report.get("queries") or []
            if record.get("status") != "completed"
        ]
        if problem_queries:
            lines += ["Problem queries:"]
            for record in problem_queries:
                lines.append(
                    f"  {record.get('query_id')} | {record.get('status')} | "
                    f"{record.get('elapsed_seconds', 0)}s | {record.get('error', '')}"
                )
            lines.append("")

    lines += ["── Programme watch ───────────────────────────────────────"]
    for label, matches in _watched_programs(new_jobs).items():
        lines.append(f"{label}: {len(matches)}")
        for job in matches:
            lines.append(f"  {job.get('company')} — {job.get('role_title')} | {job.get('url', '')}")
    lines.append("")

    lines += ["── Volume by lane ────────────────────────────────────────"]

    for lane in ("A", "B", "C"):
        lane_jobs = [j for j in new_jobs if str(j.get("lane") or "A").upper() == lane]
        lane_proceed = sum(j.get("decision") == "Proceed" for j in lane_jobs)
        lane_reject = sum(j.get("decision") in ("Reject", "Deprioritize") for j in lane_jobs)
        lane_unsure = sum(j.get("decision") == "Unsure" for j in lane_jobs)
        lane_errors = sum(j.get("decision") == "Error" for j in lane_jobs)
        lines.append(
            f"Lane {lane}: {len(lane_jobs)} found | {lane_proceed} proceed | "
            f"{lane_reject} rejected/deprioritized | {lane_unsure} unsure | {lane_errors} errors"
        )
    lines.append("")

    rejection_reasons = Counter(_rejection_reason(j) for j in rejected)
    rejection_reasons.pop("", None)
    if rejection_reasons:
        lines += ["── Rejects by reason ──────────────────────────────────────"]
        for reason, count in rejection_reasons.most_common():
            lines.append(f"{count:>3}  {reason}")
        lines.append("")

    if high:
        lines += ["── High Priority ─────────────────────────────────────────"]
        for j in high:
            lines += [
                f"  [{j.get('fit_score')}/10]  {j.get('company')} — {j.get('role_title')}",
                f"           {j.get('location')} | {j.get('source')}",
                f"           {j.get('fit_rationale', '')}",
                f"           {j.get('url', '')}",
                "",
            ]

    if mid:
        lines += ["── Medium Priority ───────────────────────────────────────"]
        for j in mid:
            lines += [
                f"  [{j.get('fit_score')}/10]  {j.get('company')} — {j.get('role_title')}",
                f"           {j.get('location')} | {j.get('fit_rationale', '')}",
                "",
            ]

    if low:
        lines += ["── Low Priority ──────────────────────────────────────────"]
        for j in low:
            lines += [
                f"  [{j.get('fit_score')}/10]  {j.get('company')} — {j.get('role_title')}",
            ]
        lines += [""]

    if rejected:
        lines += ["── Rejected / Skipped ────────────────────────────────────"]
        for j in rejected:
            lines += [
                f"  ✗  {j.get('company')} — {j.get('role_title')}",
                f"     {j.get('fit_rationale', '')}",
            ]
        lines += [""]

    if unsure:
        lines += ["── Unsure / Review ────────────────────────────────────────"]
        for j in unsure:
            lines += [
                f"  ?  {j.get('company')} — {j.get('role_title')}",
                f"     {j.get('fit_rationale', '')}",
            ]
        lines += [""]

    if errors:
        lines += ["── Scoring Errors (will retry next run) ──────────────────"]
        for j in errors:
            lines += [
                f"  ⚠  {j.get('company')} — {j.get('role_title')}",
                f"     {j.get('fit_rationale', '')}",
            ]
        lines += [""]

    lines += [f"{'='*60}", f"End of log"]

    log_path.write_text("\n".join(lines), encoding="utf-8")
    return log_path


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run(hours_old: int = 24,
        dry_run:   bool = False,
        skip_scrape: bool = False,
        model: str = "claude-haiku-4-5-20251001",
        results_override: int | None = None,
        query_timeout_seconds: int = 120,
        run_timeout_seconds: int = 5400,
        with_startup_apply: bool = False,
        startup_limit_companies: int = 12,
        startup_limit_jobs: int = 30,
        startup_sources: set[str] | None = None,
        verbose: bool = True) -> list[dict]:
    """
    Full pipeline run. Returns list of newly added + scored job dicts.
    """
    run_start = datetime.now()
    run_deadline_epoch = time.time() + max(run_timeout_seconds, 1)

    # Lazy imports — keeps startup fast and errors localised
    sys.path.insert(0, str(_HERE))
    from scraper import scrape
    from scorer  import score_batch, _load_api_key
    import anthropic

    if verbose:
        print(f"\n{'═'*60}")
        print(f"  Job Pipeline — {run_start.strftime('%Y-%m-%d %H:%M')}")
        print(f"{'═'*60}")

    # ── Step 1: Load existing jobs ────────────────────────────────────────────
    df_existing = load_jobs()
    existing_hashes = get_existing_hashes(df_existing)
    next_id = int(df_existing["id"].dropna().astype(float).max() + 1) \
              if not df_existing.empty and df_existing["id"].dropna().any() else 1

    if verbose:
        print(f"  Existing jobs in xlsx: {len(df_existing)}")

    # ── Step 2: Scrape (or load unscored 'new' jobs) ─────────────────────────
    scrape_report: dict = {}
    if skip_scrape:
        # Re-score any jobs already in xlsx with status=new and no fit_score.
        # fillna("") required — xlsx NaN cells read back as float NaN, not "".
        mask = (df_existing["status"] == "new") & (df_existing["fit_score"].fillna("") == "")
        new_jobs = df_existing[mask].to_dict("records")
        if verbose:
            print(f"  --skip-scrape: found {len(new_jobs)} unscored 'new' jobs in xlsx")
    else:
        new_jobs = scrape(
            hours_old=hours_old,
            existing_hashes=existing_hashes,
            results_override=results_override,
            verbose=verbose,
            per_query_timeout_seconds=max(query_timeout_seconds, 1),
            total_timeout_seconds=max(int(run_deadline_epoch - time.time() - 30), 1),
            run_report=scrape_report,
        )

    scored_jobs: list[dict] = []
    if not new_jobs:
        if verbose:
            print("  No new jobs to process in the standard lane.")
    else:
        # ── Step 3: Score ─────────────────────────────────────────────────────
        api_key = _load_api_key()
        client  = anthropic.Anthropic(api_key=api_key)

        if verbose:
            # Cost estimate before committing — assumes ~1,800 input + 80 output tokens/job
            # Haiku:  $0.80/M input  + $4.00/M output  ≈ $0.00176/job
            # Sonnet: $3.00/M input  + $15.00/M output ≈ $0.00660/job
            n = len(new_jobs)
            if "haiku" in model.lower():
                cost_per = 0.00176
            elif "sonnet" in model.lower():
                cost_per = 0.00660
            else:
                cost_per = 0.00500   # safe fallback for unknown model
            est_cost = n * cost_per
            print(f"\n  Scoring {n} jobs with {model}")
            print(f"  Estimated API cost: ~${est_cost:.2f}  ({n} × ${cost_per:.5f}/job)")

        scored_jobs = score_batch(
            new_jobs,
            client=client,
            model=model,
            verbose=verbose,
            deadline_epoch=max(run_deadline_epoch - 15, time.time()),
        )

        # ── Step 4: Write to xlsx ─────────────────────────────────────────────
        if not skip_scrape:
            rows   = jobs_to_rows(scored_jobs, start_id=next_id)
            df_new = pd.DataFrame(rows, columns=COLUMNS)
            df_all = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            # Update existing rows in-place
            df_all = df_existing.copy()
            for job in scored_jobs:
                mask = df_all["url_hash"] == job.get("url_hash", "")
                for col in ["fit_score", "fit_rationale", "role_type", "status"]:
                    if mask.any() and job.get(col) is not None:
                        df_all.loc[mask, col] = job[col]

        save_jobs(df_all, dry_run=dry_run)

        if verbose and not dry_run:
            print(f"\n  ✓ jobs.xlsx updated  ({len(df_all)} total rows)")

    # ── Step 5: Digest + log ──────────────────────────────────────────────────
    # Always emit a run-scoped report, including successful zero-result runs.
    print_digest(scored_jobs, run_start, scrape_report=scrape_report)

    log_path = write_run_log(
        scored_jobs,
        run_start,
        hours_old,
        dry_run,
        scrape_report=scrape_report,
    )
    if verbose:
        print(f"  Run log → {log_path}")

    startup_jobs: list[dict] = []
    if with_startup_apply:
        if verbose:
            print("\n  Running startup apply lane...")
        from startup_apply_pipeline import run as run_startup_apply

        startup_jobs = run_startup_apply(
            dry_run=dry_run,
            skip_score=False,
            model=model,
            limit_companies=startup_limit_companies,
            limit_jobs=startup_limit_jobs,
            include_sources=startup_sources,
            verbose=verbose,
        )

    return [*scored_jobs, *startup_jobs]


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Job pipeline — scrape → score → write to jobs.xlsx"
    )
    parser.add_argument(
        "--hours-old", type=int, default=24,
        help="Scraper lookback window in hours (default: 24)"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Scrape and score but don't write to jobs.xlsx"
    )
    parser.add_argument(
        "--skip-scrape", action="store_true",
        help="Skip scraping — re-score jobs already in xlsx with status=new"
    )
    parser.add_argument(
        "--model", type=str, default="claude-haiku-4-5-20251001",
        help="Claude model for scoring (default: claude-haiku-4-5-20251001)"
    )
    parser.add_argument(
        "--quiet", action="store_true",
        help="Suppress per-job verbose output"
    )
    parser.add_argument(
        "--results", type=int, default=None,
        help="Override RESULTS_WANTED per query per site (e.g. --results 200 for validation runs)"
    )
    parser.add_argument(
        "--query-timeout", type=int, default=120,
        help="Hard wall-clock timeout for each JobSpy query in seconds (default: 120)",
    )
    parser.add_argument(
        "--run-timeout", type=int, default=5400,
        help="Hard total pipeline wall-clock cap in seconds (default: 5400)",
    )
    parser.add_argument(
        "--with-startup-apply", action="store_true",
        help="Also run the startup-apply lane after the standard LinkedIn/Indeed lane"
    )
    parser.add_argument(
        "--startup-limit-companies", type=int, default=12,
        help="Maximum companies/pages to inspect per startup source (default: 12)"
    )
    parser.add_argument(
        "--startup-limit-jobs", type=int, default=30,
        help="Maximum startup jobs to keep before scoring (default: 30)"
    )
    parser.add_argument(
        "--startup-source", action="append", default=[],
        help="Optional startup source_id filter, repeatable (for example: --startup-source builtin_sf_job_lists)"
    )
    parser.add_argument("--_supervised-worker", action="store_true", help=argparse.SUPPRESS)
    args = parser.parse_args()

    if not args._supervised_worker:
        supervisor_started = datetime.now()
        command = [
            sys.executable,
            str(Path(__file__).resolve()),
            *sys.argv[1:],
            "--_supervised-worker",
        ]
        process = subprocess.Popen(
            command,
            env={**os.environ, "PYTHONUNBUFFERED": "1"},
        )
        try:
            raise SystemExit(_wait_process_wall_clock(process, max(args.run_timeout, 1) + 5))
        except subprocess.TimeoutExpired:
            process.terminate()
            try:
                _wait_process_wall_clock(process, 5)
            except subprocess.TimeoutExpired:
                process.kill()
                _wait_process_wall_clock(process, 5)
            LOGS_DIR.mkdir(exist_ok=True)
            timeout_log = LOGS_DIR / f"run_{supervisor_started.strftime('%Y-%m-%d_%H%M')}_timeout.txt"
            timeout_log.write_text(
                "\n".join(
                    [
                        "Pipeline supervisor timeout",
                        "=" * 60,
                        f"Started: {supervisor_started.isoformat(timespec='seconds')}",
                        f"Hard cap: {max(args.run_timeout, 1)}s",
                        "Action: child process terminated; completed query checkpoints remain usable",
                    ]
                ),
                encoding="utf-8",
            )
            print(
                f"\n  ✗ Pipeline hard wall-clock cap reached; worker terminated. "
                f"Timeout log → {timeout_log}",
                flush=True,
            )
            raise SystemExit(124)

    run(
        hours_old=args.hours_old,
        dry_run=args.dry_run,
        skip_scrape=args.skip_scrape,
        model=args.model,
        results_override=args.results,
        query_timeout_seconds=max(args.query_timeout, 1),
        run_timeout_seconds=max(args.run_timeout, 1),
        with_startup_apply=args.with_startup_apply,
        startup_limit_companies=max(args.startup_limit_companies, 1),
        startup_limit_jobs=max(args.startup_limit_jobs, 1),
        startup_sources={value.strip() for value in args.startup_source if value.strip()} or None,
        verbose=not args.quiet,
    )
