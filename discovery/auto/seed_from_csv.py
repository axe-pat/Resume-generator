"""
seed_from_csv.py  — Re-seed jobs.xlsx from Akshat_Master_Application_Tracker_NO_UNKNOWNS.csv

Dedup rules:
  1. Drop rows whose Job Title looks like an email confirmation / subject line
  2. Drop garbage company rows (credit card emails, supply chain Breda, etc.)
  3. Remap special rows where the company field is actually an email subject
     and the real company must be extracted (Okta, McKinsey, MGB, etc.)
  4. Normalize company names (typos, HTML entities, suffixes)
  5. Deduplicate: per canonical_company keep one row per distinct job title
     (same-ish titles → pick the richer row; genuinely different titles → keep both)
  6. Overwrite Jobs sheet in jobs.xlsx, preserving Reference sheet
"""

import hashlib, re
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
CSV_PATH   = Path("/sessions/gifted-upbeat-curie/mnt/uploads/Akshat_Master_Application_Tracker_NO_UNKNOWNS.csv")
JOBS_XLSX  = Path(__file__).parent.parent / "jobs.xlsx"   # discovery/jobs.xlsx
SHEET_NAME = "Jobs"

COLUMNS = [
    "id", "date_found", "date_posted", "company", "role_title", "role_type",
    "location", "url", "url_hash", "source",
    "fit_score", "fit_rationale", "status",
    "date_applied", "folder_path", "resume_run", "jd_text", "notes",
    "lane", "deadline", "deadline_source", "everify_status",
    "sponsorship_flag", "classification", "reject_reason",
]

# ── Email / junk title patterns ───────────────────────────────────────────────
# If a Job Title matches any of these it's an email subject, not a real title
_EMAIL_PATS = re.compile(
    r"^thank you for"
    r"|^re:\s"
    r"|^regarding your application"
    r"|your application to\b"
    r"|^security code"
    r"|^verify your"
    r"|^an update (about|from|on) your"
    r"|akshat,?\s+thank you"
    r"| at okta$"
    r"|- akshat pathak$"
    r"|^re: thank you"
    r"|akshat, thank you"
    r"|^akshat, thank"
    r"|follow up on your application"
    r"|re: stolen card"
    r"|^your application to\b"
    r"|re: google 2026 mba internship - invitation to interview"   # keep as interview signal handled below
    ,
    re.IGNORECASE,
)

# Titles that are obviously NOT job titles (notes, role type labels, etc.)
_JUNK_TITLES = re.compile(
    r"^role unspecified$"
    r"|^search pm intern role"
    r"|^mba product manager intern pm, pmm, marketing$"
    r"|^\s*$"
    ,
    re.IGNORECASE,
)

# ── Garbage company names ─────────────────────────────────────────────────────
_GARBAGE_CO_PAT = (
    r"(?i)^product manager intern(?: role)?$"
    r"|(?i)^reflect my correct name$"
    r"|(?i)^position supply chain internship open day$"
)

# ── Company name normalisation map ────────────────────────────────────────────
# Keys are lowercase stripped; values are canonical display names
COMPANY_MAP = {
    "1001 amgen inc":                          "Amgen",
    "7-eleven inc":                            "7-Eleven",
    "7-eleven":                                "7-Eleven",
    "activision blizzard king":                "Activision Blizzard",
    "apollo management holdings, l.p":         "Apollo",
    "apollo":                                  "Apollo",
    "arcesium llc":                            "Arcesium",
    "cisco":                                   "Cisco",
    "cisco ":                                  "Cisco",
    "cloudflare mba intern":                   "Cloudflare",
    "cloudflare monetization intern":          "Cloudflare",
    "databaricks":                             "Databricks",
    "gofundme":                                "GoFundMe",
    "hewlett packard enterprise":              "Hewlett Packard Enterprise",
    "hpe":                                     "Hewlett Packard Enterprise",
    "ibm":                                     "IBM",
    "ibm ":                                    "IBM",
    "mba product management intern":           "Okta",   # email body gave the company away
    "mckinsey &amp":                           "McKinsey",
    "mckinsey &amp;":                          "McKinsey",
    "mgb &lt":                                 "Mass General Brigham",
    "mongodb":                                 "MongoDB",
    "mongo db":                                "MongoDB",
    "nike, inc":                               "Nike",
    "nvidia":                                  "NVIDIA",
    "nvidia ":                                 "NVIDIA",
    "nvidiaexternalcareersite":                "NVIDIA",
    "pricewaterhousecoopers advisory services llc": "PwC",
    "us_entry_level":                          "PwC",
    "samsung research america internship":     "Samsung Research America",
    "santander (deposit)":                     "Santander",
    "santander(deposit)":                      "Santander",
    "santander":                               "Santander",
    "schneider":                               "Schneider Electric",
    "second dinner":                           "Second Dinner",
    "sigma":                                   "Sigma Computing",
    "snorkel ai":                              "Snorkel AI",
    "tiktok":                                  "TikTok",
    "typeface":                                "Typeface",
    "unavailable":                             "Unknown",
    "unknown company":                         "Unknown",
    "varkada":                                 "Verkada",
    "zoom communications":                     "Zoom",
    "zoox":                                    "Zoox",
    "a10":                                     "A10 Networks",
    "zoom":                                    "Zoom",
    "phygtl":                                  "Phygtl",
    "duolingo":                                "Duolingo",
    "glean":                                   "Glean",
}

def normalise_company(raw: str) -> str:
    key = str(raw).strip().lower()
    return COMPANY_MAP.get(key, str(raw).strip())


# ── Title canonicalisation for dedup keying ───────────────────────────────────
# Strip common junk suffixes so "Product Manager Intern (PM0003)" and
# "Product Manager Intern" are treated as the same title for dedup.
_TITLE_CLEANUP = re.compile(
    r"\s*\(PM\d+\)\s*$"      # trailing role ID
    r"|\s*\(\d{4,}\)\s*$"    # trailing numeric ID
    r"|\s*-\s*summer 202\d.*$"  # trailing date junk
    r"|\s*202\d.*$"           # trailing year
    r"|\s+intern(ship)?$"    # trailing "intern" to unify variants
    ,
    re.IGNORECASE,
)

def title_key(title: str) -> str:
    t = str(title).strip().lower()
    # Remove TikTok platform context first: "[TikTok-Product-Search Growth]- "
    t = re.sub(r"\[tiktok-[^\]]*\]\s*-?\s*", "", t)
    # Strip leading bracket/year: "[2026] " or "2026 " or "Summer 2026] "
    t = re.sub(r"^(?:\[?\d{4}\]?\s+|summer\s+\d{4}\]\s*)", "", t)
    # Strip season+year phrase ANYWHERE (e.g. "- Summer 2026" mid-title)
    t = re.sub(r"\s*[-\u2013\u2014]?\s*\(?(?:summer|fall|spring)\s*202\d\b[^)]*\)?\s*", " ", t)
    # Strip remaining standalone years "(2026)" or " 2026"
    t = re.sub(r"\s*\(?202\d\)?\s*", " ", t)
    # Strip role IDs: (PM0003), (R158879), (56279), (A67690) — up to 3 letters then digits
    t = re.sub(r"\s*\([a-z]{0,3}\d{3,}\)\s*", " ", t)
    # Strip parenthetical acronyms 2-6 chars: (ALA), (MBA), (DDV), (BS/MS)
    t = re.sub(r"\s*\([a-z/]{2,6}\)\s*", " ", t)
    # Strip long trailing parenthetical notes "(Nationwide Opportunities)"
    t = re.sub(r"\s*\([^)]{10,}\)\s*$", "", t)
    # Strip "cross-business", "multi-function", "some PM/PMM..." qualifier noise
    t = re.sub(r",?\s*cross-?business\b.*$", "", t)
    t = re.sub(r",?\s*multi-?function\b.*$", "", t)
    t = re.sub(r",?\s*some\s+pm\b.*$", "", t)
    # Strip company-name prefixes that appear in role titles
    t = re.sub(r"^amazon\s+", "", t)
    t = re.sub(r"^nike,?\s*(?:inc\.?\s*)?", "", t)
    t = re.sub(r"^nvidia\s+", "", t)
    # Normalise "product management" → "product manager"
    t = t.replace("product management", "product manager")
    # Normalise "apm intern" → "associate product manager intern"
    t = re.sub(r"\bapm\s+intern\b", "associate product manager intern", t)
    # Normalise "pm intern" → "product manager intern"
    t = re.sub(r"\bpm\s+intern\b", "product manager intern", t)
    # Strip "general pm, ..." and "mba and non-mba..." suffix noise
    t = re.sub(r",?\s*general pm.*$", "", t)
    t = re.sub(r",?\s*mba and non-mba.*$", "", t)
    t = re.sub(r"\s+pm-specific\s*$", "", t)
    # Normalise "internship(s)" → "intern"
    t = re.sub(r"\binternships?\b", "intern", t)
    # Strip standalone season words left over after year stripping (e.g. "... summer ...")
    t = re.sub(r"\b(?:summer|fall|spring|winter)\b", " ", t)
    # Normalise em-dashes / en-dashes to space
    t = re.sub(r"[\u2013\u2014]+", " ", t)
    # Remove remaining punctuation clutter (commas, brackets)
    t = re.sub(r"[,;\[\]]", " ", t)
    # Collapse whitespace
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _url_hash(url: str) -> str:
    return hashlib.md5(str(url).strip().lower().encode()).hexdigest()


# ── Status → normalised value ─────────────────────────────────────────────────
def normalise_status(raw_status: str) -> str:
    s = str(raw_status).strip().lower()
    if s in ("applied", "pending", "rejected", "interview"):
        return "applied"
    return "applied"   # default — if it's in the tracker they applied


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    df = pd.read_csv(CSV_PATH, dtype=str).fillna("")

    # Rename CSV columns to internal names
    df.columns = ["company", "app_date", "role_id", "title", "csv_status"]

    rows_raw = len(df)
    print(f"Loaded {rows_raw} rows from CSV")

    # ── Step 1: Drop garbage company rows ────────────────────────────────────
    def is_garbage_co(s):
        s = str(s).strip()
        return bool(re.match(_GARBAGE_CO_PAT, s, re.IGNORECASE))
    mask_garbage_co = df["company"].apply(is_garbage_co)
    df = df[~mask_garbage_co].copy()
    print(f"  Dropped {mask_garbage_co.sum()} garbage-company rows → {len(df)} remain")

    # ── Step 2: Identify and handle special "company is wrong" rows ───────────
    # Row: company=MGB &Lt, title="Re: Thank You For Applying to MGB"
    # → company=Mass General Brigham, title=Product Manager Intern (unknown)
    # Row: company=Mckinsey &Amp, title="Akshat, thank you..."
    # → company=McKinsey, title=MBA Consulting Intern
    # Row: company=Mba Product Management Intern, title="Thank you for applying...at Okta"
    # → company=Okta, title=MBA Product Management Intern
    for idx, row in df.iterrows():
        co  = str(row["company"]).strip().lower()
        ttl = str(row["title"]).strip()
        if co == "mgb &lt":
            df.at[idx, "company"] = "Mass General Brigham"
            df.at[idx, "title"]   = "Product Manager Intern"
        elif co in ("mckinsey &amp", "mckinsey &amp;"):
            df.at[idx, "company"] = "McKinsey"
            df.at[idx, "title"]   = "MBA Summer Associate"
        elif co == "mba product management intern":
            df.at[idx, "company"] = "Okta"
            df.at[idx, "title"]   = "MBA Product Management Intern"

    # ── Step 3: Handle "only proof of application is a thank-you email" rows ──
    # Some companies appear ONLY as a thank-you email row. We want to keep them
    # as an applied entry but fix the title.
    # Detect them: title matches email pattern.
    # We'll handle this after step 4 (email drop) — we'll catch ones that would
    # otherwise be orphaned.

    # ── Step 3b: Fix garbled / junk titles ───────────────────────────────────
    # "Summer 2026] Product Management Intern" → "Product Management Intern"
    df["title"] = df["title"].str.replace(
        r"^(?:Summer\s*)?\d{4}\]\s*", "", regex=True
    )
    # Drop purely tech-keyword non-titles like "AI platforms, GPU"
    _NOT_TITLE = re.compile(
        r"^ai platforms"
        r"|^gpu\b",
        re.IGNORECASE,
    )
    df = df[~df["title"].apply(lambda t: bool(_NOT_TITLE.match(str(t).strip())))].copy()

    # ── Step 4: Drop email-subject title rows ─────────────────────────────────
    # BUT first snapshot which companies have at least one non-email real title.
    # For companies that are ONLY email rows we'll synthesise a placeholder.
    def is_email_title(s):
        return bool(_EMAIL_PATS.search(str(s).strip()))
    def is_junk_title(s):
        return bool(_JUNK_TITLES.search(str(s).strip()))
    is_email = df["title"].apply(is_email_title)
    is_junk  = df["title"].apply(is_junk_title)
    is_bad   = is_email | is_junk

    # Companies that have at least one good title
    df["_canon_co"] = df["company"].apply(normalise_company)
    good_title_cos  = set(df.loc[~is_bad, "_canon_co"].unique())

    # For companies ONLY represented by bad-title rows, synthesise a real row
    only_email_cos = {}
    for idx, row in df[is_bad].iterrows():
        co = normalise_company(row["company"])
        if co not in good_title_cos:
            # First time we see this company with only email rows → keep one entry
            if co not in only_email_cos:
                # Extract a plausible title from the email text if possible
                ttl_raw = str(row["title"])
                # Try to pull company-implied title
                only_email_cos[co] = {
                    "company":  co,
                    "app_date": row["app_date"],
                    "title":    "Product Manager Intern",
                    "csv_status": row["csv_status"],
                }

    # Now drop all email/junk rows
    df = df[~is_bad].copy()
    print(f"  Dropped {is_bad.sum()} email-subject / junk-title rows → {len(df)} remain")

    # Append synthesised rows for email-only companies
    synth_rows = list(only_email_cos.values())
    if synth_rows:
        df_synth = pd.DataFrame(synth_rows)
        df_synth["role_id"] = ""
        df = pd.concat([df, df_synth], ignore_index=True)
        print(f"  Synthesised {len(synth_rows)} rows for email-only companies: "
              f"{list(only_email_cos.keys())}")

    # ── Step 5: Normalise company names ──────────────────────────────────────
    df["_canon_co"] = df["company"].apply(normalise_company)

    # ── Step 6: Drop Cloudflare/NVIDIA/Santander bare duplicates ─────────────
    # Some rows have the same canonical company + same rough title more than once.
    df["_title_key"] = df["title"].apply(title_key)
    df["_dedup_key"] = df["_canon_co"] + "||" + df["_title_key"]

    # Within each dedup_key, keep the row with the most information
    # Scoring heuristic: prefer row with role_id > row with earlier date > row with longer title
    def row_score(r):
        score = 0
        if str(r.get("role_id", "")).strip(): score += 10
        title_len = len(str(r.get("title", "")))
        score += min(title_len, 20)  # up to 20 bonus for descriptive title
        try:
            pd.to_datetime(r.get("app_date", ""))
            score += 5
        except Exception:
            pass
        return score

    df["_score"] = df.apply(row_score, axis=1)
    df = df.sort_values("_score", ascending=False)
    df = df.drop_duplicates(subset=["_dedup_key"], keep="first").copy()
    print(f"  After title-level dedup: {len(df)} rows")

    # ── Step 6b: Prefix dedup within each company ─────────────────────────────
    # If company has both "product manager intern" and "product manager intern – search growth",
    # the generic "product manager intern" is redundant — drop it.
    keep_idx = set(df.index)
    by_company = df.groupby("_canon_co")
    for _, grp in by_company:
        keys = grp["_dedup_key"].tolist()
        idxs = grp.index.tolist()
        for i, (ki, ii) in enumerate(zip(keys, idxs)):
            for j, (kj, ij) in enumerate(zip(keys, idxs)):
                if i != j and ki != kj and kj.startswith(ki + " "):
                    # ki is a prefix of kj → ki is redundant
                    keep_idx.discard(ii)
                    break
    dropped = len(df) - len(keep_idx)
    df = df.loc[list(keep_idx)].copy()
    if dropped:
        print(f"  Prefix dedup removed {dropped} generic rows → {len(df)} remain")

    # ── Step 7: Map to jobs.xlsx schema ──────────────────────────────────────
    today = datetime.now().strftime("%Y-%m-%d")

    output_rows = []
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        canon_co = row["_canon_co"]
        raw_title = str(row["title"]).strip()

        # Parse application date
        app_date = ""
        try:
            dt = pd.to_datetime(row["app_date"])
            if pd.notna(dt):
                app_date = dt.strftime("%Y-%m-%d")
        except Exception:
            pass

        # Placeholder URL hash
        placeholder = f"{canon_co.lower().replace(' ', '_')}_{title_key(raw_title)}"
        url_hash = hashlib.md5(placeholder.encode()).hexdigest()

        output_rows.append({
            "id":            i,
            "date_found":    app_date or today,
            "company":       canon_co,
            "role_title":    raw_title,
            "role_type":     "",
            "location":      "",
            "url":           "",
            "url_hash":      url_hash,
            "source":        "manual",
            "fit_score":     "",
            "fit_rationale": "",
            "status":        "applied",
            "date_applied":  app_date,
            "folder_path":   "",
            "jd_text":       "",
            "notes":         "",
        })

    df_out = pd.DataFrame(output_rows, columns=COLUMNS)

    # ── Step 8: Write to jobs.xlsx ────────────────────────────────────────────
    if JOBS_XLSX.exists():
        with pd.ExcelWriter(
            JOBS_XLSX, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df_out.to_excel(writer, sheet_name=SHEET_NAME, index=False)
    else:
        with pd.ExcelWriter(JOBS_XLSX, engine="openpyxl") as writer:
            df_out.to_excel(writer, sheet_name=SHEET_NAME, index=False)

    # ── Re-apply formatting ───────────────────────────────────────────────────
    wb = load_workbook(JOBS_XLSX)
    ws = wb[SHEET_NAME]

    header_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    applied_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    col_widths = {
        "id": 6, "date_found": 13, "company": 22, "role_title": 35,
        "role_type": 12, "location": 16, "url": 40, "url_hash": 14,
        "source": 10, "fit_score": 10, "fit_rationale": 50, "status": 12,
        "date_applied": 13, "folder_path": 40, "resume_run": 12, "jd_text": 18, "notes": 40,
        "lane": 8, "deadline": 22, "deadline_source": 18, "everify_status": 16,
        "sponsorship_flag": 28, "classification": 14, "reject_reason": 50,
    }

    for i, col in enumerate(df_out.columns, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 18)

    ws.freeze_panes    = "A2"
    ws.row_dimensions[1].height = 28

    # Green tint for all data rows (all are "applied")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = applied_fill

    wb.save(JOBS_XLSX)

    # ── Report ────────────────────────────────────────────────────────────────
    print(f"\n{'═'*55}")
    print(f"  Seeded jobs.xlsx with {len(df_out)} clean applied rows")
    print(f"  (started with {rows_raw} CSV rows)")
    print(f"\n  Company sample (first 10):")
    for _, r in df_out.head(10).iterrows():
        print(f"    {r['company']:28s}  {r['role_title'][:40]}")
    print(f"\n  Unique companies: {df_out['company'].nunique()}")
    print(f"{'═'*55}\n")


if __name__ == "__main__":
    main()
