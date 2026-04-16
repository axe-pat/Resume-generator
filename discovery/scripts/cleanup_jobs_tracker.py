#!/usr/bin/env python3
from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

import jobs  # noqa: E402


ARCHIVE_REASON = "[cleanup] archived legacy skip row"
STATUS_NORMALIZATION = {
    "applied": "applied",
    "Applied": "applied",
    "queued": "queued",
    "promoted": "promoted",
    "generated": "generated",
    "review": "review",
    "new": "new",
    "closed": "closed",
    "parked": "parked",
    "failed": "failed",
    "rejected": "rejected",
    "reject": "rejected",
    "skip": "skipped",
    "skipped": "skipped",
}


def _normalize_status(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    return STATUS_NORMALIZATION.get(raw, STATUS_NORMALIZATION.get(raw.lower(), raw.lower()))


def main() -> int:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with jobs.XlsxLock():
        df = jobs.load_jobs()
        df = df.copy()
        df["status"] = df["status"].apply(_normalize_status)

        skip_mask = df["status"].isin(["skip", "skipped"])
        to_archive = df[skip_mask].copy()
        active_df = df[~skip_mask].copy()

        if to_archive.empty:
            jobs.save_jobs(active_df)
            print("No skip/skipped rows found. Status normalization still applied.")
            return 0

        to_archive["notes"] = to_archive["notes"].fillna("").astype(str).apply(
            lambda notes: f"{notes}\n{ARCHIVE_REASON} at {timestamp}".strip() if notes.strip() else f"{ARCHIVE_REASON} at {timestamp}"
        )

        wb = load_workbook(jobs.JOBS_XLSX)
        if jobs.ARCHIVE_SHEET not in wb.sheetnames:
            ws_arch = wb.create_sheet(jobs.ARCHIVE_SHEET)
            ws_arch.append(jobs.COLUMNS)
        else:
            ws_arch = wb[jobs.ARCHIVE_SHEET]

        for _, row in to_archive.iterrows():
            ws_arch.append([str(row.get(col, "")) for col in jobs.COLUMNS])
        wb.save(jobs.JOBS_XLSX)
        wb.close()

        jobs.save_jobs(active_df)

    print(f"Archived skip rows: {len(to_archive)}")
    print(f"Remaining active rows: {len(active_df)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
